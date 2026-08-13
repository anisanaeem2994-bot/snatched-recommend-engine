"""
build_final_sheet.py

Builds a properly formatted pending_approval sheet for a given month,
matching the exact visual structure the owner is already used to from
August — same colors, same merged customer header rows, same dropdown,
same layout — so every month looks identical to her, regardless of what
changed underneath.

Usage:
    python3 build_final_sheet.py 2026-09 pending_approval_september

This replaces the flat/plain fallback sheet with the real formatted one.
"""
import sys
import importlib
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from collections import Counter

import recommend_v5
importlib.reload(recommend_v5)
from recommend_v5 import (set_target_month, load_customers, load_box_history,
                            is_customer_due, recommend, TARGET_MONTH)

SHEET_PATH = '/home/claude/snatched_beauty_box_master.xlsx'

def set_sheet_path(new_path):
    """Lets the web service point this at a temp file per-request, instead
    of always reading/writing the hardcoded sandbox path."""
    global SHEET_PATH
    SHEET_PATH = new_path

# ---- Exact styling confirmed from the real August sheet ----
HEADER_FILL   = PatternFill('solid', fgColor='C27BA0')   # column header row
CUST_FILL     = PatternFill('solid', fgColor='B4A7D6')   # customer name row
PRODUCT_FILL  = PatternFill('solid', fgColor='FCE5CD')   # product rows
HEADER_FONT   = Font(name='Arial', size=11, bold=True, color='FFFFFF')
CUST_FONT     = Font(name='Arial', size=11, bold=True, color='400080')
CUST_SUBFONT  = Font(name='Arial', size=9,  bold=False, color='400080')
PRODUCT_FONT  = Font(name='Arial', size=10, bold=False)
COL_WIDTHS    = {'A':4,'B':46,'C':22,'D':11,'E':7,'F':54,'G':14,'H':65,'I':14,'J':40}
COLUMN_HEADERS = ['#', 'Product Name', 'Category', 'Tier', 'Stock',
                   'Why Recommended — AI Curator Notes', 'STATUS ✏️',
                   'Owner Notes / Swap Request', 'Price (AED)',
                   '💬 Owner Comments (any time)']

def build_month_sheet(target_month, sheet_name, skip_customers=None, tier_overrides=None):
    """
    target_month: 'YYYY-MM'
    sheet_name: name of the tab to (re)create, e.g. 'pending_approval_september'
    skip_customers: optional set of customer names to exclude even if due
                    (e.g. someone who just cancelled after the pre-build)
    tier_overrides: optional dict {customer_name: 'Essentials'/'Prestige'/etc}
                    to build a box at a DIFFERENT tier than their stored
                    profile, without changing that stored profile (e.g. a
                    customer trying a different tier before committing).
    """
    skip_customers = skip_customers or set()
    tier_overrides = tier_overrides or {}
    set_target_month(target_month)

    customers = load_customers()
    all_rec, recent_boxes, box_timeline = load_box_history()

    wb = openpyxl.load_workbook(SHEET_PATH)
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    # Column widths
    for col, width in COL_WIDTHS.items():
        ws.column_dimensions[col].width = width

    # Header row
    for i, h in enumerate(COLUMN_HEADERS, start=1):
        c = ws.cell(1, i, h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(vertical='center')
    ws.row_dimensions[1].height = 21.75

    # Dropdown validation, applied per-customer-block once we know the ranges
    row = 2
    for cid, cust in customers.items():
        if cust['name'] in skip_customers:
            continue
        due, reason = is_customer_due(cust, box_timeline, target_month)
        if not due:
            continue

        override = tier_overrides.get(cust['name'])
        out = recommend(cust['name'], box_type_override=override)
        if isinstance(out, str):
            print(f'COULD NOT BUILD for {cust["name"]}: {out}')
            continue
        (cust_r, picks, warnings, received, ratio, recent, hard_block,
         soft_avoid, hist_pat, total, timeline, inv_cat_map, value_summary) = out

        # -- Customer header row --
        last_box = None
        for m, prods in reversed(box_timeline.get(cid, [])):
            last_box = m
            break
        e_count = sum(1 for p in picks if p['tier'] == 'Essentials')
        p_count = sum(1 for p in picks if p['tier'] == 'Prestige')
        age = recommend_v5.get_age(cust.get('birthday', ''))
        display_box_type = override if override else cust['box_type']
        trial_note = '  ⚠️ TRIAL — profile still shows different tier' if override else ''
        header_text = (f"  {cust['name']}   |   {display_box_type}{trial_note}  "
                        f"({e_count}E + {p_count}P)   |   Last box: {last_box}   |   "
                        f"Month: {target_month}, Age: {age}")
        ws.cell(row, 1, header_text)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row, 7, '← use dropdown in STATUS column per product below')
        ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=8)
        for c in range(1, 11):
            cell = ws.cell(row, c)
            cell.fill = CUST_FILL
            cell.font = CUST_FONT if c == 1 else CUST_SUBFONT
        ws.row_dimensions[row].height = 21.75
        header_row = row
        row += 1

        # -- Product rows --
        first_product_row = row
        for i, p in enumerate(picks, 1):
            price_str = p.get('retail_price_aed')
            why = recommend_v5.build_explanation(
                p, cid, hard_block, soft_avoid, hist_pat, total,
                timeline, inv_cat_map,
                recommend_v5.load_preferences().get(cid, {}),
                recommend_v5.load_frequencies().get(cid, {}),
                recommend_v5.load_quiz().get(cid, {}),
                age, cust['notes'])
            values = [i, p['name'], p['category'], p['tier'], p['stock'],
                      why, None, None, price_str, None]
            for c, v in enumerate(values, 1):
                cell = ws.cell(row, c, v)
                cell.fill = PRODUCT_FILL
                cell.font = PRODUCT_FONT
                if c in (6, 8, 10):
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
            ws.row_dimensions[row].height = 90
            row += 1
        last_product_row = row - 1

        # Dropdown validation for this customer's 5 status cells
        dv = DataValidation(type='list', formula1='"approved,swap,skip"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f'G{first_product_row}:G{last_product_row}')

        # -- Box balance / total summary row --
        cat_counts = Counter(p['category'].split(' - ')[0] for p in picks)
        brands = ', '.join(sorted(set(recommend_v5.get_brand(p['name']) for p in picks)))
        balance_str = '  |  '.join(f'{k}: {v}' for k, v in cat_counts.items())
        status_icon = '✅' if value_summary['met_target'] else \
                      '🟡' if value_summary.get('marginal') else '🔴'
        summary_text = (f"Box balance: {balance_str}     Brands: {brands}   |   "
                         f"Target: AED {value_summary['target']}   |   "
                         f"BOX TOTAL: {value_summary['total_retail']:.2f} AED {status_icon}")
        ws.cell(row, 1, summary_text)
        ws.row_dimensions[row].height = 13.5
        row += 1

        # Spacer row
        row += 1

        print(f'Built {cust["name"]}: {len(picks)} products, '
              f'AED {value_summary["total_retail"]} / {value_summary["target"]}')

    wb.save(SHEET_PATH)
    print(f'\nSheet "{sheet_name}" built for {target_month}.')

if __name__ == '__main__':
    target = sys.argv[1] if len(sys.argv) > 1 else '2026-09'
    sheet = sys.argv[2] if len(sys.argv) > 2 else f'pending_approval_{target}'
    build_month_sheet(target, sheet)
