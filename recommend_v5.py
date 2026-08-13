"""
Snatched Beauty Box — Recommendation Engine v5
Adds: time-gap tracking, history pattern context, all quiz fields
"""
import openpyxl, sys
from collections import defaultdict, Counter
from datetime import datetime

# ---- Target-month override ----
# By default, all "how many months ago" / "is this customer due" calculations
# use the real current date. Setting TARGET_MONTH (format 'YYYY-MM') lets the
# whole script build a box for a DIFFERENT month — e.g. pre-building September
# while it's still August — without every date calculation silently using the
# wrong "today". Call set_target_month('2026-09') before running recommend().
TARGET_MONTH = None

def set_target_month(month_str):
    """month_str format: 'YYYY-MM'. Pass None to revert to using the real
    current date (normal behavior)."""
    global TARGET_MONTH
    TARGET_MONTH = month_str

def _effective_now():
    """Returns the datetime to treat as 'today' for all month-math —
    either the real current date, or the 1st of TARGET_MONTH if set."""
    if TARGET_MONTH:
        return datetime.strptime(TARGET_MONTH + '-01', '%Y-%m-%d')
    return datetime.now()

import os
_this_dir = os.path.dirname(os.path.abspath(__file__))

def _find_bundled_workbook():
    """Looks for the bundled spreadsheet regardless of its exact filename —
    picks whichever .xlsx file is sitting alongside this script. This avoids
    needing to match one specific exact filename on GitHub."""
    for fname in os.listdir(_this_dir):
        if fname.lower().endswith('.xlsx'):
            return os.path.join(_this_dir, fname)
    return None

path = _find_bundled_workbook()
try:
    wb = openpyxl.load_workbook(path) if path else None
except FileNotFoundError:
    wb = None

def set_workbook_path(new_path):
    """Reloads the workbook from a different file — used by the web service
    so each request can run against a freshly-exported copy of the live
    Google Sheet, without changing any of the actual recommendation logic
    below. Must be called before load_customers()/recommend()/etc."""
    global wb
    wb = openpyxl.load_workbook(new_path)

_allocated_this_cycle = {}

def reset_allocations():
    global _allocated_this_cycle
    _allocated_this_cycle = {}

def load_customers():
    ws = wb['customers']
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]: continue
        out[row[0]] = {
            'id': row[0], 'name': row[1], 'email': row[2],
            'phone': row[3], 'status': row[4], 'payment': row[5],
            'box_type': row[6], 'cadence': row[7], 'city': row[8],
            'birthday': str(row[9]) if row[9] else '',
            'subscribed_since': str(row[10]) if row[10] else '',
            'notes': row[14] or '',
        }
    return out

def load_box_history():
    ws_b = wb['boxes']
    cust_boxes = defaultdict(list)
    for row in ws_b.iter_rows(min_row=2, values_only=True):
        if row[1]: cust_boxes[row[1]].append((str(row[3]), row[0]))

    for cid in cust_boxes:
        cust_boxes[cid].sort(key=lambda x: x[0])

    ws_i = wb['box_items']
    box_products = defaultdict(list)
    for row in ws_i.iter_rows(min_row=2, values_only=True):
        if not row[5]: continue
        prod = str(row[6]).lower().strip() if row[6] else ''
        if prod and 'flag:' not in prod:
            box_products[row[5]].append(prod)

    all_received    = defaultdict(set)
    recent_boxes    = defaultdict(list)
    box_timeline    = defaultdict(list)

    for cid, box_list in cust_boxes.items():
        for month, box_id in box_list:
            prods = box_products.get(box_id, [])
            all_received[cid].update(prods)
            box_timeline[cid].append((month, prods))
        last4 = box_list[-4:][::-1]
        for _, box_id in last4:
            recent_boxes[cid].append(box_products.get(box_id, []))

    return all_received, recent_boxes, box_timeline

def load_preferences():
    ws = wb['quiz_preferences']
    p = defaultdict(lambda: defaultdict(list))
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1]: p[row[1]][row[3]].append(row[4])
    return p

def load_frequencies():
    ws = wb['quiz_frequencies']
    f = defaultdict(dict)
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1]: f[row[1]][row[3]] = row[4]
    return f

def load_quiz():
    ws = wb['quiz_responses']
    q = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1]: q[row[1]] = {
            'skin_tone': row[2], 'eye_color': row[3],
            'hair_color': row[4], 'makeup_comfort': row[5]
        }
    return q

def load_inventory():
    ws = wb['inventory']
    inv = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r[0]: continue
        inv.append({
            'id': r[0], 'name': r[1], 'tier': r[2],
            'category': r[3], 'stock': r[4] or 0,
            'cost_price': r[5] if len(r) > 5 else None,
            'retail_price_aed': r[6] if len(r) > 6 else None,
            'price_source': r[7] if len(r) > 7 else None,
        })
    return inv

BOX_RATIOS = {
    'Essentials':     {'Essentials': 5, 'Prestige': 0},
    'Prestige':       {'Essentials': 0, 'Prestige': 5},
    'Mixed Glam X':   {'Essentials': 4, 'Prestige': 1},
    'Mixed Glam XX':  {'Essentials': 3, 'Prestige': 2},
    'Mixed Glam XXX': {'Essentials': 2, 'Prestige': 3},
}

BOX_VALUE_TARGETS = {
    'Essentials':     200,
    'Prestige':       600,
    'Mixed Glam X':   300,
    'Mixed Glam XX':  400,
    'Mixed Glam XXX': 500,
}

BOX_SELLING_PRICE = {
    ('Essentials', 'Monthly'):       99,
    ('Essentials', 'Bi-Monthly'):    119,
    ('Prestige', 'Monthly'):         389,
    ('Prestige', 'Bi-Monthly'):      409,
    ('Mixed Glam X', 'Monthly'):     169,
    ('Mixed Glam XX', 'Monthly'):    239,
    ('Mixed Glam XXX', 'Monthly'):   299,
}

def get_selling_price(box_type, cadence):
    return BOX_SELLING_PRICE.get((box_type, cadence),
           BOX_SELLING_PRICE.get((box_type, 'Monthly')))

def optimize_box_value(results, pool_by_tier, box_type):
    """
    Checks the box's total retail value against the target for its type.
    If below target and the gap isn't marginal, tries swapping the
    lowest-value pick(s) for higher-value eligible alternatives in the
    same tier/category from the already-filtered pool — never pulling in
    anything that failed the eligibility checks upstream. Per the stated
    policy: a small/marginal shortfall is left alone rather than forced,
    since meeting preference/variety/category rules matters more than
    exactly hitting a value number. Returns (results, value_summary dict).
    """
    target = BOX_VALUE_TARGETS.get(box_type)
    if not target:
        return results, {'target': None, 'total_retail': None, 'met_target': None}

    def total_value(items):
        return sum(p.get('retail_price_aed') or 0 for p in items)

    current_total = total_value(results)
    MARGINAL_PCT = 0.10  # within 10% of target counts as "marginal" — leave as is
    marginal_floor = target * (1 - MARGINAL_PCT)

    swapped_any = False
    if current_total < marginal_floor:
        # Try swapping cheapest picks for pricier same-category/tier alternatives
        results_sorted = sorted(results, key=lambda p: p.get('retail_price_aed') or 0)
        for i, low_item in enumerate(results_sorted):
            if total_value(results) >= target:
                break
            tier = low_item['tier']
            cat = low_item['category']
            candidates = [p for p in pool_by_tier.get(tier, [])
                          if p['category'] == cat
                          and p['name'] != low_item['name']
                          and (p.get('retail_price_aed') or 0) > (low_item.get('retail_price_aed') or 0)]
            if candidates:
                # Pick whichever candidate gets the box closest to (but not
                # wildly past) the target — not just the single most expensive
                # item available, which can badly overshoot in one swap.
                gap_needed = target - current_total
                best = min(candidates,
                           key=lambda p: abs((p.get('retail_price_aed') or 0)
                                              - (low_item.get('retail_price_aed') or 0)
                                              - gap_needed))
                idx = results.index(low_item)
                price_before = total_value(results)
                results[idx] = best
                current_total = total_value(results)
                swapped_any = True

    final_total = total_value(results)

    # NEW: handle the opposite case — a box that's significantly OVER target
    # (e.g. a single expensive item pushed the whole box past budget). Try
    # swapping the priciest pick(s) down toward a cheaper eligible alternative,
    # same "closest to target" logic as the under-target case above.
    over_ceiling = target * (1 + MARGINAL_PCT)
    if final_total > over_ceiling:
        results_sorted_desc = sorted(results, key=lambda p: -(p.get('retail_price_aed') or 0))
        for high_item in results_sorted_desc:
            if total_value(results) <= over_ceiling:
                break
            tier = high_item['tier']
            cat = high_item['category']
            candidates = [p for p in pool_by_tier.get(tier, [])
                          if p['category'] == cat
                          and p['name'] != high_item['name']
                          and (p.get('retail_price_aed') or 0) < (high_item.get('retail_price_aed') or 0)]
            if candidates:
                gap_needed = target - total_value(results)
                best = min(candidates,
                           key=lambda p: abs((p.get('retail_price_aed') or 0)
                                              - (high_item.get('retail_price_aed') or 0)
                                              - gap_needed))
                idx = results.index(high_item)
                results[idx] = best
                swapped_any = True
        final_total = total_value(results)

    met_target = final_total >= target
    gap = round(target - final_total, 2) if not met_target else 0

    summary = {
        'target': target,
        'total_retail': round(final_total, 2),
        'met_target': met_target,
        'gap': gap,
        'swapped_for_value': swapped_any,
        'marginal': (not met_target) and final_total >= marginal_floor,
    }
    return results, summary

FREQ_TO_CAT = {
    'cleanser': 'Skincare - Cleanser',
    'eye_treatment': 'Skincare - Eye Care',
    'serum_moisturizer': 'Skincare - Serum',
    'sheet_masks': 'Skincare - Mask',
    'primer': 'Face - Primer',
    'concealer': 'Face - Concealer',
    'foundation': 'Face - Foundation',
    'makeup_tools': 'Tools - Accessories',
    'powder': 'Face - Setting Powder',
    'blush': 'Face - Blush',
    'bronzer_contour': 'Face - Bronzer',
    'highlighter': 'Face - Highlighter',
    'mascara': 'Eyes - Mascara',
    'brows': 'Eyes - Brows',
    'eyeshadow': 'Eyes - Eyeshadow',
    'eyeliner': 'Eyes - Eyeliner',
    'lipstick': 'Lips - Lipstick',
    'liquid_lipstick': 'Lips - Liquid Lipstick',
    'lip_gloss': 'Lips - Gloss',
    'lip_liner': 'Lips - Liner',
}

CAT_TO_FREQ = {v: k for k, v in FREQ_TO_CAT.items()}
FREQ_WEIGHT = {'Often': 3, 'Sometimes': 2, 'Rarely': 0}
LIP_SUBCATS = {'Lips - Lipstick', 'Lips - Liquid Lipstick', 'Lips - Gloss',
               'Lips - Liner', 'Lips - Balm', 'Lips - Stain',
               'Lips - Plumper', 'Lips - Lip Oil'}

HAIR_CONCERN_KEYWORDS = {
    'Frizzy':          ['frizz', 'anti-frizz', 'smooth', 'sleek', 'taming', 'humidity'],
    'Dry and damaged': ['repair', 'damage', 'moisture', 'nourish', 'strengthen', 'bond'],
    'Needs volume':    ['volume', 'volumis', 'lift', 'thicken', 'fuller', 'body'],
    'Split Ends':      ['split', 'bond', 'repair', 'strengthen', 'seal'],
}

SCENT_KEYWORDS = {
    'Floral':         ['floral', 'rose', 'jasmine', 'lavender', 'cherry blossom', 'peony'],
    'Earthy & woody': ['woody', 'oud', 'musk', 'cedar', 'sandalwood', 'amber'],
    'Fresh':          ['fresh', 'citrus', 'mint', 'cucumber', 'aqua', 'clean'],
    'Spicy':          ['spicy', 'pepper', 'cardamom', 'warm spice'],
}

SKIN_TONE_SHADES = {
    'Fair':   ['fair', 'light', 'ivory', 'porcelain', 'alabaster'],
    'Light':  ['light', 'fair', 'natural', 'shell', 'buff'],
    'Medium': ['medium', 'natural', 'beige', 'sand', 'warm'],
    'Tan':    ['tan', 'warm', 'golden', 'bronze', 'caramel', 'honey'],
    'Dark':   ['dark', 'deep', 'rich', 'mahogany'],
    'Deep':   ['deep', 'dark', 'ebony', 'rich'],
}

def get_brand(product_name):
    known_brands = [
        'La Roche-Posay', 'Too Faced', 'Charlotte Tilbury', 'Rare Beauty',
        'Fenty Beauty', 'Fenty Skin', 'Anastasia Beverly Hills',
        'Urban Decay', 'First Aid Beauty', 'Milk Makeup', 'Makeup By Mario',
        'Makeup For Ever', 'Cover FX', 'Dr. Dennis Gross', 'Physicians Formula',
        'Wet n Wild', 'NYX Professional Makeup', 'Some By Mi',
        'Daily Life Forever', 'Drunk Elephant', 'Beauty of Joseon',
        'Soap & Glory', 'Pixi by Petra', 'e.l.f.', 'Real Techniques',
        'Masque Bar', 'NYX', 'Covergirl', 'CoverGirl', 'Maybelline',
        'Revlon', 'Rimmel', 'Bourjois', 'Essence', 'Marcelle', 'Annabelle',
        'Eveline', 'Ryshi', 'Danielle Creations', 'Andalou', 'Sukin',
        'Garnier', 'Biore', 'Eucerin', 'Smashbox', 'Nudestix', 'Tarte',
        'Huda Beauty', 'MAC', 'Lancôme', 'Giorgio Armani', 'Tom Ford',
        'NARS', 'Benefit', 'Zoeva', 'Clinique', 'Estée Lauder', 'Laneige',
        'COSRX', 'Medicube', 'Pixi', 'Sephora', 'Glossier', 'REM Beauty',
        'Stila', 'GXVE', 'Morphe', 'TIRTIR', 'Kosas', 'Laura Mercier',
        'Youngblood', 'Hourglass', 'Joah', 'Daylogic', 'Alya Skin',
    ]
    name_lower = product_name.lower()
    for brand in known_brands:
        if name_lower.startswith(brand.lower()):
            return brand.lower()
    return product_name.split()[0].lower()

# Brand name variants that should be treated as identical when comparing
# whether two historical/inventory entries are "the same brand" — since a
# product's name may or may not include a full sub-brand qualifier
# (e.g. "NYX" vs "NYX Professional Makeup") depending on how/when it was recorded.
_BRAND_ALIASES = {
    'nyx professional makeup': 'nyx',
    'covergirl': 'covergirl', 'cover girl': 'covergirl',
    'pixi by petra': 'pixi',
    'e.l.f.': 'elf',
}

# Filler/marketing words stripped out before comparing product names, since
# their presence or absence is exactly the kind of harmless wording drift
# (not a real product difference) that was breaking exact-match duplicate
# detection (e.g. "Professional Makeup", "Cosmetics", "Collection").
_FILLER_WORDS = {
    'professional', 'makeup', 'cosmetics', 'collection', 'the', 'a', 'an',
    'and', 'by', 'for', 'with', 'of', 'in',
}

def _split_name_shade(product_name):
    """Split a product name into (base_name, shade) using the same
    convention used elsewhere in this codebase (' – ' or ' (' delimiter)."""
    name = product_name.strip()
    base = name.split(' – ')[0].split(' (')[0].strip()
    shade = None
    if ' – ' in name:
        shade = name.split(' – ', 1)[1].strip()
    elif '(' in name and name.endswith(')'):
        shade = name.split('(', 1)[1].rstrip(')').strip()
    return base, shade

def _normalize_tokens(text):
    """Lowercase, strip punctuation, remove filler words, return a set of
    the remaining significant words — used so word-order differences and
    minor wording drift don't block a real duplicate match."""
    import re
    text = text.lower()
    text = re.sub(r'[^a-z0-9\s]', ' ', text)
    words = [w for w in text.split() if w and w not in _FILLER_WORDS]
    return set(words)

def products_match(name_a, name_b):
    """
    Compares two product names and returns one of:
      'exact'       — same brand, same core product, same (or no) shade —
                      this is the case that must NEVER be re-sent to a customer.
      'same_shade_variant' — same brand and core product, but a DIFFERENT
                      shade — allowed, but only after a 6-month gap.
      'different'   — genuinely different products.
    Matching is done on brand + a token-set comparison of the core name
    (ignoring word order and filler words) rather than requiring an exact
    string match, since historical records and current inventory naming
    can drift slightly (extra/missing qualifier words, punctuation,
    word order) without actually being different products.
    """
    base_a, shade_a = _split_name_shade(name_a)
    base_b, shade_b = _split_name_shade(name_b)

    brand_a = _BRAND_ALIASES.get(get_brand(name_a), get_brand(name_a))
    brand_b = _BRAND_ALIASES.get(get_brand(name_b), get_brand(name_b))
    if brand_a != brand_b:
        return 'different'

    tokens_a = _normalize_tokens(base_a) - _normalize_tokens(brand_a)
    tokens_b = _normalize_tokens(base_b) - _normalize_tokens(brand_a)
    if not tokens_a or not tokens_b:
        return 'different'

    overlap = len(tokens_a & tokens_b)
    smaller = min(len(tokens_a), len(tokens_b))
    similarity = overlap / smaller if smaller else 0

    # Require the smaller name's tokens to be almost entirely contained in
    # the larger one — handles "X" being a trimmed-down version of "X Y Z"
    if similarity < 0.75:
        return 'different'

    # A format/texture word present on only one side means these are
    # genuinely different products even if they share a family name and
    # most other words (e.g. "Soft Pinch Liquid Blush" vs "Soft Pinch
    # Luminous Powder Blush" — same family, different formula).
    _FORMAT_WORDS = {
        'liquid', 'powder', 'cream', 'matte', 'gel', 'stick', 'balm', 'oil',
        'cushion', 'foam', 'gloss', 'tint', 'wand', 'pencil', 'pen', 'spray',
        'mist', 'serum', 'essence', 'mask', 'wipes', 'pads', 'pad', 'crayon',
        'stain', 'butter', 'whip',
    }
    if (tokens_a ^ tokens_b) & _FORMAT_WORDS:
        return 'different'

    # Core product matches — now check the shade
    shade_match = (shade_a is None and shade_b is None) or \
                  (shade_a and shade_b and _normalize_tokens(shade_a) == _normalize_tokens(shade_b)) or \
                  (shade_a is None) != (shade_b is None)  # one side just omitted the shade in
                                                            # the record — treat as the same product
                                                            # rather than assume it's a different one
    return 'exact' if shade_match else 'same_shade_variant'

def guess_category_from_name(product_name):
    """
    Category fallback used when a historical product name doesn't exactly
    match current inventory. Ordered so SPECIFIC multi-word phrases that
    would otherwise collide with generic single-word keywords in other
    categories are resolved first (e.g. 'hair mask'/'hair serum' before
    generic 'mask'/'serum'; 'lip liner' before generic 'liner'; 'eye mask'
    before generic 'mask'; 'pore stamping' before generic 'pore').
    """
    name = product_name.lower()

    # ---- Step 1: cross-category collision resolution (most specific first) ----
    if any(x in name for x in ['hair mask', 'hair serum', 'scalp serum', 'scalp treatment']):
        return 'Haircare'
    if any(x in name for x in ['lip liner', 'lip pencil', 'lip definer']):
        return 'Lips - Liner'
    if any(x in name for x in ['eye mask', 'eye gel mask', 'under eye mask', 'under-eye mask',
                                'eye patch', 'under-eye patch', 'eye cream', 'eye gel',
                                'eye serum', 'eye treatment', 'de-puffer']):
        return 'Skincare - Eye Care'
    # Any clear "mask" signal is resolved here, BEFORE cleanser/primer/serum checks below —
    # otherwise products like "Cleansing Sheet Mask" or "Pore Perfect...Mask" get grabbed
    # by the word "cleansing" or "pore" before ever reaching the mask check.
    if any(x in name for x in ['mask', 'masque', 'peel off', 'peel-off', 'sheet mask']):
        return 'Skincare - Mask'
    if any(x in name for x in ['pore minimizing', 'pore refining', 'poreless', 'pore primer',
                                'pore strip']):
        return 'Face - Primer'
    if 'primer' in name:
        return 'Face - Primer'  # e.g. "Eyeshadow Primer" is fundamentally a primer, not a shadow

    # ---- Step 2: general category checks ----
    if any(x in name for x in ['mascara', 'lash']): return 'Eyes - Mascara'
    if any(x in name for x in ['eyeliner', 'eye liner', 'kohl', 'kajal']): return 'Eyes - Eyeliner'
    if any(x in name for x in ['eyeshadow', 'shadow', 'palette', 'eye shadow']): return 'Eyes - Eyeshadow'
    if any(x in name for x in ['brow', 'eyebrow']): return 'Eyes - Brows'
    if any(x in name for x in ['lip gloss', 'gloss', 'lip shine']): return 'Lips - Gloss'
    if any(x in name for x in ['liquid lip', 'lip cream', 'lip colour', 'lip color', 'lip stain', 'lip tint']):
        return 'Lips - Liquid Lipstick'
    if any(x in name for x in ['lipstick', 'lip stick', 'lip butter', 'lip balm']): return 'Lips - Lipstick'
    if any(x in name for x in ['lip oil', 'lip drip']): return 'Lips - Lip Oil'
    if 'liner' in name: return 'Eyes - Eyeliner'  # bare fallback, lip liner already resolved above
    if any(x in name for x in ['blush', 'blusher', 'cheek']): return 'Face - Blush'
    if any(x in name for x in ['bronzer', 'bronze', 'bronzing', 'contour']): return 'Face - Bronzer'
    if any(x in name for x in ['highlighter', 'highlight', 'glow', 'luminous', 'illuminat']): return 'Face - Highlighter'
    if any(x in name for x in ['foundation', 'base', 'bb cream', 'cc cream']): return 'Face - Foundation'
    if any(x in name for x in ['concealer', 'conceale']): return 'Face - Concealer'
    if any(x in name for x in ['powder', 'setting powder', 'pressed powder']): return 'Face - Setting Powder'
    if 'base coat' in name: return 'Face - Primer'
    if any(x in name for x in ['cleanser', 'cleansing', 'face wash', 'micellar', 'cleanse']): return 'Skincare - Cleanser'
    if any(x in name for x in ['serum', 'essence', 'ampoule']): return 'Skincare - Serum'
    if any(x in name for x in ['moisturis', 'moisturiz', 'cream', 'lotion', 'day cream', 'night cream']):
        return 'Skincare - Moisturizer'
    if any(x in name for x in ['toner', 'tonic', 'mist', 'spray', 'fix']): return 'Skincare - Toner'
    if any(x in name for x in ['hair', 'shampoo', 'conditioner', 'scalp']): return 'Haircare'
    if any(x in name for x in ['body', 'scrub', 'body wash', 'bath', 'hand cream', 'lotion']): return 'Body Care'
    if any(x in name for x in ['brush', 'sponge', 'tool', 'applicator', 'roller', 'gua sha']): return 'Tools - Accessories'
    if any(x in name for x in ['nail', 'cuticle']): return 'Nail Care'
    return None

def is_customer_due(cust, box_timeline, target_month_str):
    """
    Checks whether a customer should get a box built for target_month_str
    ('YYYY-MM'). Returns (is_due: bool, reason: str).
    Catches: cancelled/paused customers, and Bi-Monthly customers who
    aren't actually due this cycle based on their real box history.
    """
    status = str(cust.get('status', '')).strip().lower()
    if status != 'active':
        return False, f"Customer status is '{cust.get('status')}', not active — no box should be built."

    cadence = str(cust.get('cadence', '')).strip().lower()
    cid = cust['id']
    history = box_timeline.get(cid, [])

    if not history:
        return True, "No prior box history — treat as due (first box)."

    last_month_str = history[-1][0]
    gap = months_since(last_month_str)
    target_gap = months_since(last_month_str)
    try:
        target_dt = datetime.strptime(target_month_str + '-01', '%Y-%m-%d')
        last_dt   = datetime.strptime(last_month_str + '-01', '%Y-%m-%d')
        months_between = (target_dt.year - last_dt.year) * 12 + (target_dt.month - last_dt.month)
    except:
        return True, "Could not parse dates — defaulting to due (verify manually)."

    if 'bi-month' in cadence or 'bimonth' in cadence:
        if months_between < 2:
            return False, f"Bi-Monthly customer — last box was {last_month_str}, only {months_between} month(s) ago. Not due until 2+ months have passed."
        return True, f"Bi-Monthly customer — last box was {last_month_str}, {months_between} months ago. Due."

    # Default: Monthly cadence — due every month customer is active
    return True, f"Monthly customer — last box was {last_month_str}, {months_between} month(s) ago. Due."

def get_age(birthday_str):
    try:
        bday = datetime.strptime(str(birthday_str)[:10], '%Y-%m-%d')
        today = datetime.now()
        return today.year - bday.year - (
            (today.month, today.day) < (bday.month, bday.day))
    except:
        return None

def months_since(month_str):
    try:
        past = datetime.strptime(month_str + '-01', '%Y-%m-%d')
        now  = _effective_now()
        return (now.year - past.year) * 12 + (now.month - past.month)
    except:
        return None

def get_category_time_gap(cid, category, box_timeline, inv_cat_map):
    for month_str, prods in reversed(box_timeline.get(cid, [])):
        for prod_lower in prods:
            cat = inv_cat_map.get(prod_lower)
            if cat is None:
                cat = guess_category_from_name(prod_lower)
            if cat == category:
                gap = months_since(month_str)
                return gap, month_str
    return None, None

def get_history_pattern(cid, box_timeline, inv_cat_map):
    pattern = Counter()
    total_products = 0
    for _, prods in box_timeline.get(cid, []):
        for prod_lower in prods:
            cat = inv_cat_map.get(prod_lower)
            if cat is None:
                cat = guess_category_from_name(prod_lower)
            if cat:
                grp = cat.split(' - ')[0]
                pattern[grp] += 1
                total_products += 1
    return pattern, total_products

def get_recent_subcats(recent_boxes, inv_cat_map):
    box_subcats = []
    for box_prods in recent_boxes:
        subcats = set()
        for prod_lower in box_prods:
            cat = inv_cat_map.get(prod_lower)
            if cat: subcats.add(cat)
        box_subcats.append(subcats)

    hard_block = set()
    if len(box_subcats) >= 3:
        hard_block = box_subcats[0] & box_subcats[1] & box_subcats[2]
    elif len(box_subcats) == 2:
        hard_block = box_subcats[0] & box_subcats[1]

    soft_avoid = set()
    for s in box_subcats:
        soft_avoid.update(s)
    soft_avoid -= hard_block

    return hard_block, soft_avoid

def check_frequency_eligibility(cat, product_name, freq, months_since_last):
    """
    Enforces the owner's stated timing rules as a hard gate (not just a
    soft score penalty):
      Often     -> min 2 months gap, EXCEPT mascara (see below)
      Sometimes -> min 3 months gap
      Rarely    -> no hard minimum here (handled by existing soft penalty
                   in score_product — Rarely items should only surface
                   once Often/Sometimes options are exhausted, which is
                   a selection-priority question, not a timing gate)
    Mascara override: black/no-shade mascara needs 4 months; a mascara
    naming a non-black color (brown/purple/green/etc.) only needs 3 months.
    Returns (is_eligible: bool, reason: str). months_since_last=None means
    never received — always eligible.
    """
    if months_since_last is None:
        return True, ""

    name = product_name.lower()
    if cat == 'Eyes - Mascara':
        colors = ['brown', 'purple', 'green', 'blue', 'plum', 'burgundy', 'navy']
        is_colored = any(c in name for c in colors)
        min_gap = 3 if is_colored else 4
    elif freq == 'Often':
        min_gap = 2
    elif freq == 'Sometimes':
        min_gap = 3
    else:
        return True, ""  # Rarely — no hard gate, just lower scoring priority

    if months_since_last < min_gap:
        return False, (f"{cat} given {months_since_last} month(s) ago — "
                        f"needs at least {min_gap} months for this frequency/category")
    return True, ""

def score_product(p, prefs, freqs, quiz, notes, hard_block, soft_avoid, age):
    score = 0
    cat   = p['category']
    name  = p['name'].lower()
    month = _effective_now().month

    if cat in hard_block: score -= 25
    if cat in soft_avoid: score -= 10

    for fk, fc in FREQ_TO_CAT.items():
        if cat == fc:
            freq = freqs.get(fk, 'Sometimes')
            score += FREQ_WEIGHT.get(freq, 1) * 3
            if freq == 'Rarely': score -= 8

    shade_map = {
        'Eyes - Eyeshadow': 'eyeshadow_color',
        'Lips - Lipstick': 'lips_color', 'Lips - Liquid Lipstick': 'lips_color',
        'Lips - Gloss': 'lips_color', 'Lips - Liner': 'lips_color',
        'Lips - Balm': 'lips_color', 'Lips - Stain': 'lips_color',
        'Lips - Plumper': 'lips_color', 'Lips - Lip Oil': 'lips_color',
        'Face - Blush': 'blush_color', 'Nail Care': 'nail_color',
    }
    pk = shade_map.get(cat)
    if pk:
        sp = [x.lower() for x in prefs.get(pk, [])]
        for shade in ['reds', 'berries', 'neutrals', 'plums', 'pinks', 'peach',
                      'mauves', 'adventurous', 'dark', 'pastels', 'smokey']:
            if shade in sp and shade in name: score += 6

    skin_tone = quiz.get('skin_tone', '')
    if skin_tone and cat in ['Face - Foundation', 'Face - Concealer',
                              'Face - Bronzer', 'Face - Setting Powder']:
        shade_words = SKIN_TONE_SHADES.get(skin_tone, [])
        if any(sw in name for sw in shade_words): score += 5
        elif any(sw in name for sw in ['fair', 'light', 'medium', 'tan',
                                        'dark', 'deep', 'ivory', 'beige']):
            score -= 4

    eye_color = quiz.get('eye_color', '').lower()
    if cat == 'Eyes - Eyeshadow':
        if 'hazel' in eye_color or 'green' in eye_color:
            if any(x in name for x in ['plum', 'purple', 'bronze', 'warm', 'copper']): score += 3
        if 'blue' in eye_color:
            if any(x in name for x in ['warm', 'bronze', 'neutral', 'copper']): score += 3

    if cat in ['Haircare', 'Skincare - Hair']:
        hair_concerns = prefs.get('hair_concerns', [])
        for concern, keywords in HAIR_CONCERN_KEYWORDS.items():
            if concern in hair_concerns:
                if any(kw in name for kw in keywords): score += 8
        if hair_concerns: score += 2

    if cat in ['Body Care', 'Skincare - Mask', 'Haircare']:
        scent_prefs = prefs.get('scent_preference', [])
        for scent, keywords in SCENT_KEYWORDS.items():
            if scent in scent_prefs:
                if any(kw in name for kw in keywords): score += 5

    sc = ' '.join(prefs.get('skin_concerns', [])).lower()
    if 'oiliness' in sc and any(x in name for x in ['oil', 'pore', 'matte', 'control', 'sebum']): score += 4
    if 'dryness'  in sc and any(x in name for x in ['hydrat', 'moisture', 'repair', 'nourish', 'barrier']): score += 4
    if 'wrinkles' in sc and any(x in name for x in ['retinol', 'anti-age', 'collagen', 'peptide', 'lift']): score += 5
    if 'acne'     in sc and any(x in name for x in ['acne', 'blemish', 'tea tree', 'salicyl', 'clarify', 'bha']): score += 5
    if 'hyper-pigmentation' in sc and any(x in name for x in ['vitamin c', 'brightening', 'niacinamide', 'turmeric']): score += 5
    if 'redness'  in sc and any(x in name for x in ['calming', 'soothing', 'centella', 'sensitive', 'aloe']): score += 5

    comfort = quiz.get('makeup_comfort', '')
    if 'not very' in str(comfort).lower():
        if any(x in name for x in ['stick', 'balm', 'tint', 'cushion', 'gloss']): score += 4
        if any(x in cat.lower() for x in ['palette', 'eyeshadow']): score -= 5
        if 'eyeliner' in cat.lower(): score -= 3
    if 'very comfortable' in str(comfort).lower():
        if any(x in cat.lower() for x in ['palette', 'liner']): score += 2

    if age:
        if age >= 40:
            if any(x in name for x in ['retinol', 'collagen', 'peptide', 'anti-age', 'lift', 'firm', 'renewal']): score += 5
            if any(x in cat for x in ['Skincare', 'Eye Care']): score += 2
        elif age <= 25:
            if any(x in cat for x in ['Lips', 'Face - Blush', 'Face - Highlighter']): score += 2

    if month in [6, 7, 8, 9]:
        if any(x in name for x in ['spf', 'sun', 'mist', 'refresh', 'light', 'water', 'cooling']): score += 3
        if any(x in name for x in ['heavy', 'rich cream']): score -= 2
    if month in [11, 12, 1, 2]:
        if any(x in name for x in ['nourish', 'rich', 'repair', 'butter', 'intensive']): score += 3

    notes_lower = notes.lower()
    if 'no nail polish' in notes_lower and 'nail' in cat.lower(): score -= 1000
    if 'no elf complexion' in notes_lower:
        if (name.startswith('elf') or name.startswith('e.l.f')) and cat in [
            'Face - Concealer', 'Face - Foundation', 'Face - Setting Powder']: score -= 1000
    if 'prefers elf' in notes_lower:
        if name.startswith('elf') or name.startswith('e.l.f'): score += 4

    return score

def pick_with_variety(scored_pool, needed, already_used=None, already_brands=None):
    MAX_PER_GROUP = 2
    picked        = []
    used_subcats  = list(already_used or [])
    used_groups   = [c.split(' - ')[0] for c in used_subcats]
    used_brands   = list(already_brands or [])

    def try_add(p, lip_limit=1, ignore_brand=False):
        cat   = p['category']
        grp   = cat.split(' - ')[0]
        brand = get_brand(p['name'])
        if cat in used_subcats: return False
        if used_groups.count(grp) >= MAX_PER_GROUP: return False
        lip_count = sum(1 for c in used_subcats if c in LIP_SUBCATS)
        if cat in LIP_SUBCATS and lip_count >= lip_limit: return False
        if not ignore_brand and brand in used_brands: return False
        picked.append(p)
        used_subcats.append(cat)
        used_groups.append(grp)
        used_brands.append(brand)
        return True

    for p in scored_pool:
        if len(picked) >= needed: break
        try_add(p, lip_limit=1)

    if len(picked) < needed:
        for p in scored_pool:
            if len(picked) >= needed: break
            if p not in picked: try_add(p, lip_limit=2)

    if len(picked) < needed:
        for p in scored_pool:
            if len(picked) >= needed: break
            if p not in picked: try_add(p, lip_limit=2, ignore_brand=True)

    if len(picked) < needed:
        for p in scored_pool:
            if len(picked) >= needed: break
            if p not in picked: picked.append(p)

    return picked[:needed]

def get_received_with_dates(cid, box_timeline):
    """Returns [(product_name, month_str), ...] for everything a customer
    has ever received — used so the shade-variant 6-month rule can check
    *when* a similar product was last sent, not just whether it was ever sent."""
    out = []
    for month_str, prods in box_timeline.get(cid, []):
        for p in prods:
            out.append((p, month_str))
    return out

def check_product_eligibility(candidate_name, received_with_dates, target_month_str):
    """
    Returns (is_eligible: bool, reason: str).
    Uses products_match() instead of exact-string comparison, so naming
    drift between historical records and current inventory (extra/missing
    qualifier words, punctuation, word order) doesn't let a real duplicate
    slip through. An 'exact' match blocks the product outright — the same
    specific item is never sent twice. A 'same_shade_variant' match is
    allowed only once 6 months have passed since that variant was sent.
    """
    try:
        target_dt = datetime.strptime(target_month_str + '-01', '%Y-%m-%d')
    except:
        target_dt = None

    for received_name, received_month in received_with_dates:
        result = products_match(candidate_name, received_name)
        if result == 'exact':
            return False, f"Same product already received ({received_month}): {received_name}"
        if result == 'same_shade_variant' and target_dt:
            try:
                r_dt = datetime.strptime(received_month + '-01', '%Y-%m-%d')
                months_between = (target_dt.year - r_dt.year) * 12 + (target_dt.month - r_dt.month)
                if months_between < 6:
                    return False, (f"Different shade of a product received {months_between} "
                                    f"month(s) ago ({received_month}: {received_name}) — "
                                    f"needs a 6-month gap between shade variants")
            except:
                pass
    return True, ""

def get_next_swap_alternative(cid, category, tier, already_rejected_names,
                                current_box_names, target_month_str, notes=''):
    """
    Finds the next-best alternative product for a category/tier the owner
    rejected — with no explanation needed from her. Reuses the exact same
    eligibility rules as the main box-building logic (never-repeat-exact,
    6-month shade-variant rule, frequency-timing gaps) PLUS excludes:
      - anything already rejected this swap round (so she's never shown
        the same item twice while cycling through options)
      - anything already elsewhere in this customer's current box
    Returns (product_dict or None, reason_if_none).
    """
    customers = load_customers()
    cust = customers.get(cid)
    if not cust:
        return None, "Customer not found"

    inventory = load_inventory()
    inv_cat_map = {p['name'].lower(): p['category'] for p in inventory}
    all_rec, recent_boxes, box_timeline = load_box_history()
    received_with_dates = get_received_with_dates(cid, box_timeline)
    freqs = load_frequencies().get(cid, {})

    rejected_lower = {n.lower().strip() for n in already_rejected_names}
    in_box_lower = {n.lower().strip() for n in current_box_names}

    candidates = []
    for p in inventory:
        if p['category'] != category or p['tier'] != tier:
            continue
        if not p.get('stock') or p['stock'] <= 0:
            continue
        if not p.get('retail_price_aed'):
            continue
        nm = p['name'].lower().strip()
        if nm in rejected_lower or nm in in_box_lower:
            continue
        if not check_product_eligibility(p['name'], received_with_dates, target_month_str)[0]:
            continue
        freq_key = next((fk for fk, fc in FREQ_TO_CAT.items() if fc == category), '')
        gap, _ = get_category_time_gap(cid, category, box_timeline, inv_cat_map)
        if not check_frequency_eligibility(category, p['name'], freqs.get(freq_key, 'Sometimes'), gap)[0]:
            continue
        if 'no nail' in notes.lower() and 'nail' in category.lower():
            continue
        if ('no elf complexion' in notes.lower() or 'no elf blush' in notes.lower()):
            if nm.startswith('elf') and category in ['Face - Concealer', 'Face - Foundation',
                                                        'Face - Setting Powder', 'Face - Blush']:
                continue
        candidates.append(p)

    if not candidates:
        return None, (f"No more eligible {category} options left in {tier} for this customer — "
                       f"every remaining product either fails a timing/duplicate rule, is out of "
                       f"stock, or was already tried this round.")

    candidates.sort(key=lambda p: -(p.get('stock') or 0))
    return candidates[0], ""

def get_next_alternative(customer_name, category, tier, already_rejected,
                          current_box_products=None):
    """
    Returns the next-best eligible replacement for a single rejected product
    — used when the owner just types "swap" with no explanation, and wants
    another option in the same slot, repeatedly, until she's satisfied.

    already_rejected: list of product names she's already said no to for
                       THIS slot this cycle — never re-suggested.
    current_box_products: the other 4 products already in her box this
                       month, so we don't introduce a same-brand-twice
                       conflict or an accidental duplicate within the box.

    Reuses all the same eligibility checks as a full recommend() run —
    duplicate detection, frequency timing, stock — just scoped to one
    category/tier instead of building a whole box from scratch.

    Returns (product_dict_or_None, message).
    """
    customers = load_customers()
    all_rec, recent_boxes, box_timeline = load_box_history()
    inventory = load_inventory()
    inv_cat_map = {p['name'].lower(): p['category'] for p in inventory}

    cust = next((c for c in customers.values()
                 if customer_name.lower() in c['name'].lower()), None)
    if not cust:
        return None, f"Customer '{customer_name}' not found."

    cid = cust['id']
    received_with_dates = get_received_with_dates(cid, box_timeline)
    recent = recent_boxes.get(cid, [])
    prefs = load_preferences().get(cid, {})
    freqs = load_frequencies().get(cid, {})
    quiz = load_quiz().get(cid, {})
    age = get_age(cust.get('birthday', ''))
    hard_block, soft_avoid = get_recent_subcats(recent, inv_cat_map)

    target_month = TARGET_MONTH if TARGET_MONTH else datetime.now().strftime('%Y-%m')
    already_rejected_lower = {r.lower().strip() for r in (already_rejected or [])}
    current_box_products = current_box_products or []
    current_brands = {get_brand(p) for p in current_box_products}

    pool = []
    for p in inventory:
        if p['category'] != category or p['tier'] != tier:
            continue
        if (p['stock'] - _allocated_this_cycle.get(p['name'].lower(), 0)) < 1:
            continue
        if p['name'].lower().strip() in already_rejected_lower:
            continue
        if not check_product_eligibility(p['name'], received_with_dates, target_month)[0]:
            continue
        freq_key = next((fk for fk, fc in FREQ_TO_CAT.items() if fc == category), '')
        if not check_frequency_eligibility(
                category, p['name'], freqs.get(freq_key, 'Sometimes'),
                get_category_time_gap(cid, category, box_timeline, inv_cat_map)[0])[0]:
            continue
        if get_brand(p['name']) in current_brands:
            continue  # no same-brand-twice-in-one-box
        pool.append(p)

    if not pool:
        return None, ("No more eligible alternatives found in this category/tier — "
                       "every option is either out of stock, already rejected, too "
                       "soon, or conflicts with another brand already in this box.")

    scored = [(score_product(p, prefs, freqs, quiz, cust['notes'], hard_block, soft_avoid, age), p)
              for p in pool]
    scored.sort(key=lambda x: -x[0])
    best = scored[0][1]
    return best, f"Next alternative: {best['name']} (AED {best.get('retail_price_aed')})"

def recommend(customer_name, box_type_override=None):
    customers  = load_customers()
    all_rec, recent_boxes, box_timeline = load_box_history()
    prefs_all  = load_preferences()
    freqs_all  = load_frequencies()
    quiz_all   = load_quiz()
    inventory  = load_inventory()

    inv_cat_map = {p['name'].lower(): p['category'] for p in inventory}

    cust = next((c for c in customers.values()
                 if customer_name.lower() in c['name'].lower()), None)
    if not cust:
        return f"Customer '{customer_name}' not found."

    cid      = cust['id']
    box_type = box_type_override if box_type_override else cust['box_type']
    notes    = cust['notes']
    received = all_rec.get(cid, set())
    received_with_dates = get_received_with_dates(cid, box_timeline)
    recent   = recent_boxes.get(cid, [])
    timeline = box_timeline
    prefs    = prefs_all.get(cid, {})
    freqs    = freqs_all.get(cid, {})
    quiz     = quiz_all.get(cid, {})
    ratio    = BOX_RATIOS.get(box_type)
    age      = get_age(cust.get('birthday', ''))

    if not ratio: return f"Unknown box type: {box_type}"

    target_month = TARGET_MONTH if TARGET_MONTH else datetime.now().strftime('%Y-%m')
    is_due, due_reason = is_customer_due(cust, timeline, target_month)
    if not is_due:
        return f"SKIP {cust['name']}: {due_reason}"

    hard_block, soft_avoid = get_recent_subcats(recent, inv_cat_map)
    history_pattern, total_products = get_history_pattern(cid, timeline, inv_cat_map)

    results, warnings = [], []
    used_subcats_global = []
    used_brands_global  = []
    pool_by_tier = {}

    for tier in ['Essentials', 'Prestige']:
        needed = ratio[tier]
        if needed == 0: continue

        pool = [p for p in inventory
                if p['tier'] == tier
                and (p['stock'] - _allocated_this_cycle.get(p['name'].lower(), 0)) >= 1
                and check_product_eligibility(p['name'], received_with_dates, target_month)[0]
                and check_frequency_eligibility(
                        p['category'], p['name'],
                        freqs.get(next((fk for fk,fc in FREQ_TO_CAT.items() if fc==p['category']), ''), 'Sometimes'),
                        get_category_time_gap(cid, p['category'], timeline, inv_cat_map)[0]
                    )[0]
                and not ('no nail' in notes.lower() and 'nail' in p['category'].lower())
                and not ('no elf complexion' in notes.lower()
                         and (p['name'].lower().startswith('elf') or
                              p['name'].lower().startswith('e.l.f'))
                         and p['category'] in ['Face - Concealer',
                             'Face - Foundation', 'Face - Setting Powder'])]
        pool_by_tier[tier] = pool

        if len(pool) < needed:
            warnings.append(f"⚠️ Only {len(pool)} eligible {tier} products (need {needed})")

        scored = sorted(pool,
                        key=lambda p: score_product(p, prefs, freqs, quiz,
                                                    notes, hard_block, soft_avoid, age),
                        reverse=True)

        picked = pick_with_variety(scored, needed,
                                   already_used=used_subcats_global,
                                   already_brands=used_brands_global)

        for p in picked:
            key = p['name'].lower()
            _allocated_this_cycle[key] = _allocated_this_cycle.get(key, 0) + 1
            used_subcats_global.append(p['category'])
            used_brands_global.append(get_brand(p['name']))

        results.extend(picked)

    results, value_summary = optimize_box_value(results, pool_by_tier, box_type)
    value_summary['selling_price'] = get_selling_price(box_type, cust.get('cadence', 'Monthly'))
    if not value_summary['met_target'] and not value_summary.get('marginal'):
        warnings.append(f"⚠️ Box retail value AED {value_summary['total_retail']} is below the "
                         f"AED {value_summary['target']} target for {box_type} "
                         f"(short by AED {value_summary['gap']}) — not enough higher-value "
                         f"eligible stock to close the gap without breaking other rules.")

    return (cust, results, warnings, received, ratio,
            recent, hard_block, soft_avoid, history_pattern,
            total_products, box_timeline, inv_cat_map, value_summary)

def build_explanation(p, cid, hard_block, soft_avoid, history_pattern,
                      total_products, box_timeline, inv_cat_map,
                      prefs, freqs, quiz, age, notes):
    cat    = p['category']
    name   = p['name'].lower()
    fkey   = CAT_TO_FREQ.get(cat)
    freq   = freqs.get(fkey, 'Sometimes') if fkey else 'Sometimes'
    grp    = cat.split(' - ')[0]
    sentences = []

    cat_readable = cat.split(' - ')[-1].lower()
    if freq == 'Often':
        sentences.append(f"She rates {cat_readable} as a frequent preference (Often).")
    elif freq == 'Sometimes':
        sentences.append(f"She has a moderate interest in {cat_readable} (Sometimes).")
    elif freq == 'Rarely':
        sentences.append(f"She doesn't prioritise {cat_readable} often, but options in her preferred categories are limited this month.")

    months_ago, last_month_str = get_category_time_gap(
        cid, cat, box_timeline, inv_cat_map)
    if months_ago is None:
        sentences.append(f"She has never received a {cat_readable} product before — this would be a first.")
    elif months_ago == 0:
        sentences.append(f"She received a {cat_readable} very recently, but options are limited this month.")
    elif months_ago == 1:
        sentences.append(f"Her last {cat_readable} was 1 month ago.")
    else:
        sentences.append(f"She hasn't received a {cat_readable} in {months_ago} months (last was {last_month_str}).")

    if total_products > 0:
        grp_count   = history_pattern.get(grp, 0)
        grp_pct     = round((grp_count / total_products) * 100)
        all_groups  = sorted(history_pattern.items(), key=lambda x: -x[1])
        top_group   = all_groups[0][0] if all_groups else ''
        if grp_count == 0:
            sentences.append(f"{grp} products have never featured in her boxes — adding variety to her history.")
        elif grp == top_group and grp_pct > 35:
            sentences.append(f"{grp} is already the most common category in her history ({grp_pct}% of all products sent) — a different category was considered, but preferences and stock guided this choice.")
        elif grp_pct < 10 and total_products > 10:
            sentences.append(f"{grp} is underrepresented in her history (only {grp_count} of {total_products} products sent) — a good opportunity to add variety.")

    sc = ' '.join(prefs.get('skin_concerns', [])).lower()
    hair_concerns = prefs.get('hair_concerns', [])
    if 'wrinkles' in sc and any(x in name for x in ['retinol', 'collagen', 'peptide', 'lift']):
        sentences.append("Matches her skin concern: wrinkles & fine lines.")
    elif 'acne' in sc and any(x in name for x in ['acne', 'blemish', 'tea tree', 'clarify']):
        sentences.append("Addresses her acne and blemish concerns.")
    elif 'dryness' in sc and any(x in name for x in ['hydrat', 'moisture', 'repair', 'barrier']):
        sentences.append("Targets her dryness and dehydration concern.")
    elif 'oiliness' in sc and any(x in name for x in ['oil', 'pore', 'matte', 'control']):
        sentences.append("Suited to her oily skin concern.")
    elif 'hyper-pigmentation' in sc and any(x in name for x in ['vitamin c', 'brightening', 'niacinamide']):
        sentences.append("Helps with her hyperpigmentation concern.")
    elif 'redness' in sc and any(x in name for x in ['calming', 'soothing', 'centella', 'sensitive']):
        sentences.append("Calming formula suited to her redness concern.")

    if cat in ['Haircare', 'Skincare - Hair'] and hair_concerns:
        for concern, keywords in HAIR_CONCERN_KEYWORDS.items():
            if concern in hair_concerns and any(kw in name for kw in keywords):
                sentences.append(f"Specifically chosen for her {concern.lower()} hair concern.")
                break

    scent_prefs = prefs.get('scent_preference', [])
    if cat == 'Body Care' and scent_prefs:
        for scent, keywords in SCENT_KEYWORDS.items():
            if scent in scent_prefs and any(kw in name for kw in keywords):
                sentences.append(f"Aligns with her {scent.lower()} scent preference.")
                break

    if cat in hard_block:
        sentences.append("This category hasn't appeared in her last 3+ boxes, making it a well-timed addition.")
    elif cat in soft_avoid:
        sentences.append("This category was in a recent box, but her frequency preference and limited options make it appropriate here.")

    comfort = quiz.get('makeup_comfort', '')
    if 'not very' in str(comfort).lower():
        if any(x in name for x in ['stick', 'balm', 'tint', 'cushion', 'gloss']):
            sentences.append("Easy-to-use format — ideal for her beginner comfort level.")

    if age and age >= 40:
        if any(x in name for x in ['retinol', 'collagen', 'peptide', 'anti-age', 'lift']):
            sentences.append(f"An age-appropriate anti-aging pick for a customer in her {(age//10)*10}s.")

    if 'prefers elf' in notes.lower() and (name.startswith('elf') or name.startswith('e.l.f')):
        sentences.append("Customer-requested brand preference.")

    return ' '.join(sentences) if sentences else "Good match for this customer's overall quiz profile and box history."

if __name__ == '__main__':
    name = ' '.join(sys.argv[1:]) if len(sys.argv) > 1 else 'Hiba'
    out  = recommend(name)
    if isinstance(out, str):
        print(out); sys.exit(1)
    cust, picks, warnings, received, ratio, recent, hard_block, soft_avoid, hist_pat, total, timeline, inv_cat_map, value_summary = out
    prefs = load_preferences().get(cust['id'], {})
    freqs = load_frequencies().get(cust['id'], {})
    quiz  = load_quiz().get(cust['id'], {})
    age   = get_age(cust.get('birthday', ''))
    print(f"\n{'='*65}")
    print(f"RECOMMENDATION: {cust['name']} | {cust['box_type']} | Age: {age}")
    print(f"{'='*65}")
    for i, p in enumerate(picks, 1):
        expl = build_explanation(p, cust['id'], hard_block, soft_avoid,
                                 hist_pat, total, timeline, inv_cat_map,
                                 prefs, freqs, quiz, age, cust['notes'])
        price_str = f"AED {p['retail_price_aed']}" if p.get('retail_price_aed') else "AED ?"
        print(f"\n{i}. {p['name']}")
        print(f"   {p['category']} | Stock: {p['stock']} | Retail: {price_str}")
        print(f"   → {expl}")
    print(f"\n{'-'*65}")
    print(f"BOX VALUE SUMMARY")
    print(f"  Paid price:        AED {value_summary['selling_price']}")
    print(f"  Total retail value: AED {value_summary['total_retail']} "
          f"(target: AED {value_summary['target']})")
    status = "✅ MEETS TARGET" if value_summary['met_target'] else \
             "🟡 BELOW TARGET (marginal, within 10%)" if value_summary.get('marginal') else \
             "🔴 BELOW TARGET"
    print(f"  Status: {status}")
    if value_summary.get('swapped_for_value'):
        print(f"  (Auto-swapped one or more items to increase retail value)")
    print(f"{'-'*65}")
    if warnings:
        for w in warnings: print(f"\n{w}")
