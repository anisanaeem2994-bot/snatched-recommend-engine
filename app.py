"""
Snatched Beauty Box — Web service wrapper for recommend_v5.py

This does NOT contain any recommendation logic itself — it just receives
requests from Make.com, saves the uploaded spreadsheet, calls the existing
(already tested) recommend_v5.py functions, and returns the result as JSON.

Endpoints:
  POST /recommend
    form-data:
      file: the current .xlsx spreadsheet (Make.com exports this fresh
            from Google Sheets before each call, so the engine always
            works with real, current data)
      customer_name: e.g. "Sandra Hoffmann"
      target_month: optional, e.g. "2026-09" — omit to use today's date
      box_type_override: optional, e.g. "Essentials"
    returns: JSON with the recommended products, value summary, warnings

  POST /swap
    form-data:
      file: the current .xlsx spreadsheet
      customer_name, category, tier: which slot is being swapped
      already_rejected: comma-separated list of product names already tried
      current_box_products: comma-separated list of the other 4 products
                             already in the box (avoids same-brand repeats)
    returns: JSON with the next alternative, or a clear "none left" message
"""
from flask import Flask, request, jsonify, session, redirect, url_for
import os
import tempfile
import importlib
import recommend_v5
from functools import wraps

import requests
import base64 as base64_lib
import threading

GITHUB_TOKEN = os.environ.get('GITHUB_TOKEN')
GITHUB_REPO = 'anisanaeem2994-bot/snatched-recommend-engine'
GITHUB_FILE_PATH = 'snatched_beauty_box_master.xlsx'

def push_to_github():
    """Called automatically after any real change (approve, unapprove,
    finalize) — immediately saves the current state to GitHub, so it
    survives even if Render restarts a moment later. Uses Python's own
    reliable requests + base64 libraries directly, not Make.com's fragile
    binary handling (which corrupted a file earlier today)."""
    if not GITHUB_TOKEN:
        print('WARNING: GITHUB_TOKEN not set — changes will NOT survive a restart.', flush=True)
        return False
    try:
        headers = {'Authorization': f'token {GITHUB_TOKEN}'}
        get_url = f'https://api.github.com/repos/{GITHUB_REPO}/contents/{GITHUB_FILE_PATH}'
        current = requests.get(get_url, headers=headers, timeout=15)
        sha = current.json().get('sha') if current.status_code == 200 else None

        with open(recommend_v5.path, 'rb') as f:
            content_b64 = base64_lib.b64encode(f.read()).decode('utf-8')

        payload = {'message': 'Auto-save after approval/change', 'content': content_b64}
        if sha:
            payload['sha'] = sha

        resp = requests.put(get_url, headers=headers, json=payload, timeout=30)
        if resp.status_code in (200, 201):
            print('Successfully pushed changes to GitHub.', flush=True)
            return True
        else:
            print(f'GitHub push failed: {resp.status_code} {resp.text[:200]}', flush=True)
            return False
    except Exception as e:
        print(f'GitHub push error: {e}', flush=True)
        return False

def push_to_github_async():
    """Fire-and-forget wrapper around push_to_github(). The real function
    makes two real network calls to GitHub's API (a GET then a PUT), each
    with its own timeout (15s + 30s) -- if GitHub is slow, or the token
    is bad and every retry/redirect eats time, that's up to ~45+ seconds
    the dashboard would otherwise sit there waiting on before the person
    using it sees ANYTHING happen, even though their change already saved
    successfully to local disk moments earlier. Every write endpoint below
    already calls wb.save(recommend_v5.path) synchronously first -- the
    actual data is safe on disk before this even starts -- so there's no
    reason the browser needs to wait on the GitHub backup part too. This
    runs it on a background thread instead, so the response comes back
    immediately regardless of how long (or whether) GitHub cooperates."""
    threading.Thread(target=push_to_github, daemon=True).start()

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 25 * 1024 * 1024

# ---- Dashboard auth (single shared password — this is a 2-person internal
# tool, not a public product, so one password kept in Render's env vars is
# enough; nothing fancier is needed) ----
app.secret_key = os.environ.get('DASHBOARD_SECRET_KEY', 'dev-only-change-me')
DASHBOARD_PASSWORD = os.environ.get('DASHBOARD_PASSWORD')
from datetime import timedelta as _timedelta, datetime as _datetime
app.permanent_session_lifetime = _timedelta(days=30)

DASHBOARD_LOGIN_HTML = """<!DOCTYPE html>
<html><head><meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1">
<title>Snatched Beauty Box — Login</title>
<style>
  body{ font-family:-apple-system,Segoe UI,Roboto,sans-serif; background:linear-gradient(135deg,#C27BA0,#B4A7D6); min-height:100vh; margin:0; display:flex; align-items:center; justify-content:center; }
  .box{ background:#fff; border-radius:16px; padding:36px 32px; width:300px; box-shadow:0 10px 30px rgba(0,0,0,0.15); text-align:center; }
  h1{ font-size:18px; margin:0 0 20px; color:#3a2f38; }
  input{ width:100%; box-sizing:border-box; padding:11px 14px; border-radius:8px; border:1px solid #ecdfe6; font-size:14px; margin-bottom:12px; }
  button{ width:100%; padding:11px; border:none; border-radius:8px; background:#C27BA0; color:#fff; font-weight:700; font-size:14px; cursor:pointer; }
  button:hover{ background:#a85f85; }
  .err{ color:#c0506a; font-size:12.5px; margin-bottom:10px; font-weight:600; }
</style></head>
<body>
  <div class="box">
    <h1>💄 Snatched Beauty Box</h1>
    <!--ERROR-->
    <form method="POST" action="/dashboard/do_login">
      <input type="password" name="password" placeholder="Password" autofocus required>
      <button type="submit">Log in</button>
    </form>
  </div>
</body></html>"""

def dashboard_login_required(f):
    """Protects every /dashboard* route. GET requests for the page itself
    get redirected to the login screen; API calls (fetch from the page's
    own JS) get a clean 401 JSON instead of an HTML redirect, since the
    frontend needs to detect 'not logged in' programmatically."""
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get('dashboard_authed'):
            if request.method == 'GET' and request.accept_mimetypes.accept_html:
                return redirect(url_for('dashboard_login_page'))
            return jsonify({'error': 'Not logged in.'}), 401
        return f(*args, **kwargs)
    return wrapper

@app.after_request
def add_cors_headers(response):
    # Lets the customer-lookup dashboard webpage call this service directly
    # from the browser (a different "origin" than Render itself).
    response.headers['Access-Control-Allow-Origin'] = '*'
    response.headers['Access-Control-Allow-Methods'] = 'GET, POST, OPTIONS'
    response.headers['Access-Control-Allow-Headers'] = 'Content-Type'
    return response  # 25MB — the exported
                                                        # spreadsheet with all
                                                        # sheets can be a few MB;
                                                        # Flask's default limit
                                                        # is much smaller and was
                                                        # silently rejecting it
                                                        # (413 error) before this.

@app.route('/', methods=['GET'])
def health_check():
    # Simple endpoint to confirm the service is alive — Make.com or a
    # browser can hit this to check the deployment worked.
    return jsonify({'status': 'ok', 'service': 'snatched-recommend-engine'})

@app.route('/dashboard/login', methods=['GET'])
def dashboard_login_page():
    """Simple one-field password screen — no username, since this is a
    2-person internal tool. On success, sets a session cookie and sends
    her straight into the dashboard."""
    if session.get('dashboard_authed'):
        return redirect('/dashboard')
    return DASHBOARD_LOGIN_HTML

@app.route('/dashboard/do_login', methods=['POST'])
def dashboard_do_login():
    if not DASHBOARD_PASSWORD:
        return jsonify({'error': 'Dashboard password is not configured on the server yet.'}), 500
    submitted = request.form.get('password', '')
    if submitted == DASHBOARD_PASSWORD:
        session['dashboard_authed'] = True
        session.permanent = True
        return redirect('/dashboard')
    return DASHBOARD_LOGIN_HTML.replace(
        '<!--ERROR-->', '<div class="err">Wrong password — try again.</div>'
    )

@app.route('/dashboard/logout', methods=['POST'])
def dashboard_logout():
    session.pop('dashboard_authed', None)
    return redirect('/dashboard/login')

_DASHBOARD_HTML_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'dashboard.html')

@app.route('/dashboard', methods=['GET'])
@dashboard_login_required
def dashboard_page():
    with open(_DASHBOARD_HTML_PATH, encoding='utf-8') as f:
        return f.read()

@app.route('/customer_lookup', methods=['GET'])
@dashboard_login_required
def customer_lookup_endpoint():
    """Powers the customer-details dashboard. Given a customer's name,
    returns everything Iqra would otherwise have to hunt across multiple
    sheets for: full product history, skin type/preferences, quiz answers,
    and frequency settings — reusing the same tested data-loading logic
    the recommendation engine itself relies on."""
    customer_name = request.args.get('name')
    if not customer_name:
        return jsonify({'error': 'name is required, e.g. ?name=Sandra Hoffmann'}), 400

    try:
        importlib.reload(recommend_v5)
        customers = recommend_v5.load_customers()
        cust = next((c for c in customers.values() if c['name'].lower() == customer_name.lower()), None)
        if not cust:
            return jsonify({'error': f'No customer found named "{customer_name}".'}), 404
        cid = cust['id']

        all_rec, recent_boxes, box_timeline = recommend_v5.load_box_history()
        # box_timeline[cid] is a list of (month, [product names]) tuples — one
        # entry per box she actually received. all_received[cid] is just a
        # deduped SET of product name strings with no month attached, so it
        # must not be used here (indexing into a string byproduct gave garbled
        # single-character "months"/"products" before this fix).
        timeline = sorted(box_timeline.get(cid, []), key=lambda x: x[0] or '')
        product_history = [
            {'product': prod, 'month': month}
            for month, prods in timeline
            for prod in prods
        ]

        quiz = recommend_v5.load_quiz().get(cid, {})
        prefs = recommend_v5.load_preferences().get(cid, {})
        freqs = recommend_v5.load_frequencies().get(cid, {})

        # Every box ever recorded for her, newest first — including
        # Cancelled ones, so the drawer can show "Cancel & restock" next
        # to any committed-but-undelivered box (e.g. one built ahead of
        # travel that turned out not to be needed) and show already-
        # cancelled ones as a record. load_box_history() above skips
        # Cancelled boxes on purpose (they must not count as received
        # history), so this reads the boxes sheet directly instead.
        ws_b = recommend_v5.wb['boxes']
        customer_boxes = [
            {'box_id': row[0], 'month': str(row[3]) if row[3] else None, 'status': row[4], 'box_type': row[5]}
            for row in ws_b.iter_rows(min_row=2, values_only=True) if row[1] == cid
        ]
        customer_boxes.sort(key=lambda b: b['month'] or '', reverse=True)

        # Exact confirmed/rejected shades on file for her, if Iqra has
        # ever logged one — read-only reference, shown only when it
        # actually exists for this customer.
        shade_notes = recommend_v5.load_shade_reference().get(cust['name'].strip().lower(), [])

        # Brands blocked for her in every category (see /dashboard/block_brand).
        ws_bb = recommend_v5.wb['blocked_brands'] if 'blocked_brands' in recommend_v5.wb.sheetnames else None
        blocked_brands = []
        if ws_bb is not None:
            for row in ws_bb.iter_rows(min_row=2, values_only=True):
                if row[0] == cid and row[1]:
                    blocked_brands.append({'brand': row[1], 'reason': row[2] or '', 'source_month': row[3] or ''})

        return jsonify({
            'customer_name': cust['name'],
            'customer_id': cid,
            'status': cust.get('status'),
            'email': cust.get('email'),
            'phone': cust.get('phone'),
            'city': cust.get('city'),
            'payment_type': cust.get('payment'),
            'box_type': cust.get('box_type'),
            # NOTE: load_customers() stores this under the key 'cadence' (that's
            # also the exact key is_customer_due() reads for bi-monthly timing --
            # do not rename it there). 'billing_cadence' below is just this
            # endpoint's own output field name for the dashboard to display/edit.
            'billing_cadence': cust.get('cadence'),
            'subscribed_since': cust.get('subscribed_since'),
            'birthday': cust.get('birthday'),
            'notes': cust.get('notes'),
            'boxes': customer_boxes[:8],
            'shade_notes': shade_notes,
            'blocked_brands': blocked_brands,
            'product_history': product_history,
            'quiz': quiz,
            'preferences': prefs,
            'frequencies': freqs,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/customer_search_product', methods=['GET'])
@dashboard_login_required
def customer_search_product_endpoint():
    """Checks whether a specific customer has ever received a specific
    product before — the 'has she had this already' search Iqra wants."""
    customer_name = request.args.get('name')
    product_query = request.args.get('product', '').lower()
    if not customer_name or not product_query:
        return jsonify({'error': 'both name and product are required.'}), 400

    try:
        importlib.reload(recommend_v5)
        customers = recommend_v5.load_customers()
        cust = next((c for c in customers.values() if c['name'].lower() == customer_name.lower()), None)
        if not cust:
            return jsonify({'error': f'No customer found named "{customer_name}".'}), 404
        cid = cust['id']

        all_rec, recent_boxes, box_timeline = recommend_v5.load_box_history()
        timeline = sorted(box_timeline.get(cid, []), key=lambda x: x[0] or '')
        matches = [
            {'product': prod, 'month': month}
            for month, prods in timeline
            for prod in prods
            if product_query in prod.lower()
        ]
        return jsonify({'customer_name': cust['name'], 'query': product_query, 'matches': matches, 'has_received': len(matches) > 0})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/dashboard/set_customer_status', methods=['POST'])
@dashboard_login_required
def dashboard_set_customer_status():
    """Lets Iqra mark a customer Cancelled (or reactivate her) straight
    from the dashboard. This is the ONLY thing that controls whether
    future months generate a box for her — is_customer_due() in
    recommend_v5 already refuses to build a box for anyone whose status
    isn't 'active', so flipping this cell here is the whole fix for
    'don't build her a box next month.' It does NOT touch any box
    that's already been committed — cancel those separately below."""
    customer_id = request.form.get('customer_id')
    new_status = request.form.get('status')
    if not customer_id or new_status not in ('Active', 'Cancelled'):
        return jsonify({'error': 'customer_id and status (Active or Cancelled) are required.'}), 400

    try:
        importlib.reload(recommend_v5)
        wb = recommend_v5.wb
        ws_c = wb['customers']
        id_col = 1
        status_col = 5  # customers sheet: id, name, email, phone, status, ...
        updated = False
        for row in ws_c.iter_rows(min_row=2):
            if row[id_col - 1].value == customer_id:
                row[status_col - 1].value = new_status
                updated = True
                break
        if not updated:
            return jsonify({'error': f'No customer found with id "{customer_id}".'}), 404

        wb.save(recommend_v5.path)
        push_to_github_async()
        return jsonify({'success': True, 'customer_id': customer_id, 'status': new_status})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/dashboard/cancel_box', methods=['POST'])
@dashboard_login_required
def dashboard_cancel_box():
    """Undoes an already-committed box that turns out isn't actually
    going out (e.g. a box built ahead of time before the owner traveled,
    for a customer who then cancels before it ships). Restocks every
    product in that box back into inventory and marks the box Cancelled
    — the row is kept (not deleted) as a record, and load_box_history()
    already skips Cancelled boxes so it won't count toward her 'received
    before' history or bi-monthly gap math."""
    box_id = request.form.get('box_id')
    if not box_id:
        return jsonify({'error': 'box_id is required.'}), 400

    try:
        importlib.reload(recommend_v5)
        wb = recommend_v5.wb
        ws_b = wb['boxes']
        ws_i = wb['box_items']

        box_row = None
        for row in ws_b.iter_rows(min_row=2):
            if row[0].value == box_id:
                box_row = row
                break
        if not box_row:
            return jsonify({'error': f'No box found with id "{box_id}".'}), 404
        if str(box_row[4].value).strip().lower() == 'cancelled':
            return jsonify({'success': True, 'already_cancelled': True})

        products = [
            row[6].value for row in ws_i.iter_rows(min_row=2)
            if row[5].value == box_id and row[6].value
        ]

        ws_inv = wb['inventory']
        hi = [ws_inv.cell(1, c).value for c in range(1, ws_inv.max_column + 1)]
        n_col = hi.index('name') + 1
        s_col = hi.index('stock_qty') + 1
        inv_rows_by_name = {}
        for row in ws_inv.iter_rows(min_row=2):
            nm = row[n_col - 1].value
            if nm:
                inv_rows_by_name[nm] = row

        restocked = []
        not_found = []
        for prod in products:
            row = inv_rows_by_name.get(prod)
            if row:
                row[s_col - 1].value = (row[s_col - 1].value or 0) + 1
                restocked.append(prod)
            else:
                not_found.append(prod)

        box_row[4].value = 'Cancelled'
        wb.save(recommend_v5.path)
        push_to_github_async()
        return jsonify({
            'success': True,
            'box_id': box_id,
            'restocked_count': len(restocked),
            'restocked_products': restocked,
            'not_found_in_inventory': not_found,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

BLOCKED_BRANDS_SHEET = 'blocked_brands'
BLOCKED_BRANDS_HEADERS = ['customer_id', 'brand', 'reason', 'source_month', 'customer_name']

def _get_blocked_brands_sheet(create=False):
    wb = recommend_v5.wb
    if BLOCKED_BRANDS_SHEET not in wb.sheetnames:
        if not create:
            return None
        ws = wb.create_sheet(BLOCKED_BRANDS_SHEET)
        ws.append(BLOCKED_BRANDS_HEADERS)
        return ws
    return wb[BLOCKED_BRANDS_SHEET]

@app.route('/dashboard/block_brand', methods=['POST'])
@dashboard_login_required
def dashboard_block_brand():
    """Lets Iqra permanently block a whole brand for one customer, in
    EVERY category — e.g. a customer messages her privately saying 'no
    Elf products at all'. Stored in its own sheet (created on first use)
    rather than buried in free-text notes, so it's unambiguous and the
    engine can actually enforce it — recommend_v5's load_blocked_brands()
    is checked by both a fresh box build and every swap path."""
    customer_id = request.form.get('customer_id')
    customer_name = request.form.get('customer_name', '')
    brand = request.form.get('brand', '').strip()
    reason = request.form.get('reason', '')
    if not customer_id or not brand:
        return jsonify({'error': 'customer_id and brand are required.'}), 400

    try:
        importlib.reload(recommend_v5)
        ws = _get_blocked_brands_sheet(create=True)
        brand_lower = brand.lower()
        # Don't add a duplicate row if this brand's already blocked for her.
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == customer_id and str(row[1] or '').strip().lower() == brand_lower:
                return jsonify({'success': True, 'already_blocked': True})

        target_month = recommend_v5.TARGET_MONTH or ''
        ws.append([customer_id, brand, reason, target_month, customer_name])
        recommend_v5.wb.save(recommend_v5.path)
        push_to_github_async()
        return jsonify({'success': True, 'brand': brand})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/dashboard/unblock_brand', methods=['POST'])
@dashboard_login_required
def dashboard_unblock_brand():
    """Removes a brand block for a customer (she's changed her mind, or
    it was added by mistake)."""
    customer_id = request.form.get('customer_id')
    brand = request.form.get('brand', '').strip().lower()
    if not customer_id or not brand:
        return jsonify({'error': 'customer_id and brand are required.'}), 400

    try:
        importlib.reload(recommend_v5)
        ws = _get_blocked_brands_sheet(create=False)
        if ws is None:
            return jsonify({'success': True, 'removed': 0})

        rows_to_delete = [
            r for r in range(ws.max_row, 1, -1)
            if ws.cell(r, 1).value == customer_id
            and str(ws.cell(r, 2).value or '').strip().lower() == brand
        ]
        for r in rows_to_delete:
            ws.delete_rows(r, 1)

        recommend_v5.wb.save(recommend_v5.path)
        push_to_github_async()
        return jsonify({'success': True, 'removed': len(rows_to_delete)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/dashboard/set_customer_notes', methods=['POST'])
@dashboard_login_required
def dashboard_set_customer_notes():
    """Freeform notes box for anything that doesn't need to be a
    structured rule — general context the owner wants on record for a
    customer. (For 'never recommend this brand', use /dashboard/block_brand
    instead — that's the one the engine actually reads and enforces.)"""
    customer_id = request.form.get('customer_id')
    notes = request.form.get('notes', '')
    if not customer_id:
        return jsonify({'error': 'customer_id is required.'}), 400

    try:
        importlib.reload(recommend_v5)
        wb = recommend_v5.wb
        ws_c = wb['customers']
        notes_col = 15  # customers sheet: ... upgrade_date, notes
        updated = False
        for row in ws_c.iter_rows(min_row=2):
            if row[0].value == customer_id:
                row[notes_col - 1].value = notes
                updated = True
                break
        if not updated:
            return jsonify({'error': f'No customer found with id "{customer_id}".'}), 404

        wb.save(recommend_v5.path)
        push_to_github_async()
        return jsonify({'success': True, 'notes': notes})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# customers sheet column numbers (1-based) for every field Iqra is allowed
# to fill in or correct herself. name/email/phone/city/birthday cover gaps
# an automated intake email couldn't fill; payment_type/box_type/billing_cadence
# cover the exact ambiguity flagged for Essentials/Prestige signups (Mixed
# tiers are always Monthly, so those never need asking).
CUSTOMER_EDITABLE_FIELDS = {
    'name': 2, 'email': 3, 'phone': 4, 'payment_type': 6, 'box_type': 7,
    'billing_cadence': 8, 'city': 9, 'birthday': 10, 'subscribed_since': 11,
    'first_box_month': 12,
}

@app.route('/dashboard/edit_customer', methods=['POST'])
@dashboard_login_required
def dashboard_edit_customer():
    """Lets Iqra fill in or correct any of a customer's core details --
    e.g. an automated intake email didn't mention Monthly vs Bi-Monthly,
    or a phone/city was missing or wrong. Only touches fields actually
    present in the request, so a partial edit never wipes out fields it
    wasn't given. Whatever gets saved here is the exact same live file
    the recommend engine reads from, so it's picked up automatically the
    next time a box is generated for her -- no separate sync step."""
    customer_id = request.form.get('customer_id')
    if not customer_id:
        return jsonify({'error': 'customer_id is required.'}), 400

    try:
        importlib.reload(recommend_v5)
        wb = recommend_v5.wb
        ws_c = wb['customers']
        target_row = None
        for row in ws_c.iter_rows(min_row=2):
            if row[0].value == customer_id:
                target_row = row
                break
        if not target_row:
            return jsonify({'error': f'No customer found with id "{customer_id}".'}), 404

        updated_fields = {}
        for field, col in CUSTOMER_EDITABLE_FIELDS.items():
            if field in request.form:
                value = request.form.get(field)
                target_row[col - 1].value = value
                updated_fields[field] = value

        if not updated_fields:
            return jsonify({'error': 'No editable fields were sent.'}), 400

        wb.save(recommend_v5.path)
        push_to_github_async()
        return jsonify({'success': True, 'customer_id': customer_id, 'updated': updated_fields})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/dashboard/edit_preferences', methods=['POST'])
@dashboard_login_required
def dashboard_edit_preferences():
    """Lets Iqra correct a customer's quiz answers, category frequencies
    (Often/Sometimes/Rarely), and style preferences straight from the
    dashboard -- for whenever she learns something new about a customer
    outside of any automated quiz email, or a resubmission comes in and
    she'd rather just retype the answers herself than trust an
    auto-overwrite. Same guarantee as every other edit endpoint: this
    writes into the one live file the engine reads from, so it's picked
    up automatically the next time a box is built for her -- no separate
    sync step, no waiting."""
    customer_id = request.form.get('customer_id')
    customer_name = request.form.get('customer_name', '')
    if not customer_id:
        return jsonify({'error': 'customer_id is required.'}), 400

    try:
        importlib.reload(recommend_v5)
        wb = recommend_v5.wb
        updated = {}

        # -- quiz_responses: one row per customer (skin_tone/eye_color/
        # hair_color/makeup_comfort). If she's never had a row here yet,
        # a new one is created rather than silently doing nothing.
        quiz_cols = {'skin_tone': 4, 'eye_color': 5, 'hair_color': 6, 'makeup_comfort': 7}
        sent_quiz = {f: request.form.get(f) for f in quiz_cols if f in request.form}
        if sent_quiz:
            ws_q = wb['quiz_responses']
            target = None
            for row in ws_q.iter_rows(min_row=2):
                if row[1].value == customer_id:
                    target = row  # last match wins -- matches load_quiz()'s own behavior
            if target is None:
                new_row = _next_append_row(ws_q)
                ws_q.cell(new_row, 1).value = f'QZ-{customer_id}-manual'
                ws_q.cell(new_row, 2).value = customer_id
                ws_q.cell(new_row, 3).value = customer_name
                for f, v in sent_quiz.items():
                    ws_q.cell(new_row, quiz_cols[f]).value = v
            else:
                for f, v in sent_quiz.items():
                    target[quiz_cols[f] - 1].value = v
            updated['quiz'] = sent_quiz

        # -- quiz_frequencies: one row per (customer_id, category). Send
        # as freq_<category>=<Often|Sometimes|Rarely> form fields.
        freq_updates = {k[5:]: v for k, v in request.form.items() if k.startswith('freq_') and v}
        if freq_updates:
            ws_f = wb['quiz_frequencies']
            found_cats = set()
            for row in ws_f.iter_rows(min_row=2):
                cid, cat = row[1].value, row[3].value
                if cid == customer_id and cat in freq_updates:
                    row[4].value = freq_updates[cat]
                    found_cats.add(cat)
            for cat, level in freq_updates.items():
                if cat not in found_cats:
                    new_row = _next_append_row(ws_f)
                    ws_f.cell(new_row, 2).value = customer_id
                    ws_f.cell(new_row, 3).value = customer_name
                    ws_f.cell(new_row, 4).value = cat
                    ws_f.cell(new_row, 5).value = level
            updated['frequencies'] = freq_updates

        # -- quiz_preferences: MULTIPLE rows can share the same
        # (customer_id, key) -- e.g. a multi-select "shopping style"
        # question. Send as pref_<key>=<comma, separated, values> and
        # this replaces the FULL set of values under that key.
        pref_updates = {k[5:]: v for k, v in request.form.items() if k.startswith('pref_')}
        if pref_updates:
            ws_p = wb['quiz_preferences']
            keys_sent = set(pref_updates.keys())
            rows_to_delete = [
                r for r in range(ws_p.max_row, 1, -1)
                if ws_p.cell(r, 2).value == customer_id and ws_p.cell(r, 4).value in keys_sent
            ]
            for r in rows_to_delete:
                ws_p.delete_rows(r, 1)
            for key, raw_values in pref_updates.items():
                for v in [x.strip() for x in raw_values.split(',') if x.strip()]:
                    new_row = _next_append_row(ws_p)
                    ws_p.cell(new_row, 2).value = customer_id
                    ws_p.cell(new_row, 3).value = customer_name
                    ws_p.cell(new_row, 4).value = key
                    ws_p.cell(new_row, 5).value = v
            updated['preferences'] = pref_updates

        if not updated:
            return jsonify({'error': 'No preference fields were sent.'}), 400

        wb.save(recommend_v5.path)
        push_to_github_async()
        return jsonify({'success': True, 'customer_id': customer_id, 'updated': updated})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/recommend', methods=['POST'])
def recommend_endpoint():
    if 'file' not in request.files:
        return jsonify({'error': 'No spreadsheet file was sent.'}), 400
    customer_name = request.form.get('customer_name')
    if not customer_name:
        return jsonify({'error': 'customer_name is required.'}), 400

    target_month = request.form.get('target_month') or None
    box_type_override = request.form.get('box_type_override') or None

    uploaded = request.files['file']
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        uploaded.save(tmp.name)
        tmp_path = tmp.name

    try:
        importlib.reload(recommend_v5)
        recommend_v5.set_workbook_path(tmp_path)
        recommend_v5.set_target_month(target_month)

        out = recommend_v5.recommend(customer_name, box_type_override=box_type_override)
        if isinstance(out, str):
            return jsonify({'error': out}), 404

        (cust, picks, warnings, received, ratio, recent, hard_block, soft_avoid,
         hist_pat, total, timeline, inv_cat_map, value_summary) = out

        return jsonify({
            'customer_name': cust['name'],
            'box_type': cust['box_type'],
            'products': [
                {
                    'name': p['name'],
                    'category': p['category'],
                    'tier': p['tier'],
                    'stock': p['stock'],
                    'price_aed': p.get('retail_price_aed'),
                }
                for p in picks
            ],
            'value_summary': value_summary,
            'warnings': warnings,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        os.unlink(tmp_path)

@app.route('/approve_item', methods=['POST'])
def approve_item_endpoint():
    """Called the moment a row gets marked 'approved'. Immediately
    decrements that product's stock in the bundled file (and saves it to
    disk) so every other customer's swap check from this point on sees
    the real, reduced number — preventing the same last-1-in-stock item
    from being offered to two different customers before either is
    actually finalized."""
    row_number = request.form.get('row_number')
    if not row_number:
        return jsonify({'error': 'row_number is required.'}), 400
    row_number = int(float(row_number))
    sheet_name = request.form.get('sheet_name', 'pending_approval')

    try:
        importlib.reload(recommend_v5)
        ws = recommend_v5.wb[sheet_name]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        name_col = headers.index('Product Name') + 1
        product_name = ws.cell(row_number, name_col).value
        if not product_name:
            return jsonify({'error': f'No product found in row {row_number}.'}), 400

        ws_inv = recommend_v5.wb['inventory']
        hi = [ws_inv.cell(1, c).value for c in range(1, ws_inv.max_column + 1)]
        n_col = hi.index('name') + 1
        s_col = hi.index('stock_qty') + 1
        updated = False
        for row in ws_inv.iter_rows(min_row=2):
            if row[n_col-1].value == product_name:
                old_stock = row[s_col-1].value or 0
                row[s_col-1].value = max(0, old_stock - 1)
                updated = True
                new_stock = row[s_col-1].value
                break

        if not updated:
            return jsonify({'error': f'Product "{product_name}" not found in inventory.'}), 400

        recommend_v5.wb.save(recommend_v5.path)
        push_to_github_async()
        return jsonify({'success': True, 'product': product_name, 'stock_before': old_stock, 'stock_after': new_stock})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/unapprove_item', methods=['POST'])
def unapprove_item_endpoint():
    """Reverses /approve_item — used if the owner changes her mind after
    approving something. Increments that product's stock back by 1."""
    row_number = request.form.get('row_number')
    if not row_number:
        return jsonify({'error': 'row_number is required.'}), 400
    row_number = int(float(row_number))
    sheet_name = request.form.get('sheet_name', 'pending_approval')

    try:
        importlib.reload(recommend_v5)
        ws = recommend_v5.wb[sheet_name]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        name_col = headers.index('Product Name') + 1
        product_name = ws.cell(row_number, name_col).value
        if not product_name:
            return jsonify({'error': f'No product found in row {row_number}.'}), 400

        ws_inv = recommend_v5.wb['inventory']
        hi = [ws_inv.cell(1, c).value for c in range(1, ws_inv.max_column + 1)]
        n_col = hi.index('name') + 1
        s_col = hi.index('stock_qty') + 1
        updated = False
        for row in ws_inv.iter_rows(min_row=2):
            if row[n_col-1].value == product_name:
                old_stock = row[s_col-1].value or 0
                row[s_col-1].value = old_stock + 1
                updated = True
                new_stock = row[s_col-1].value
                break

        if not updated:
            return jsonify({'error': f'Product "{product_name}" not found in inventory.'}), 400

        recommend_v5.wb.save(recommend_v5.path)
        push_to_github_async()
        return jsonify({'success': True, 'product': product_name, 'stock_before': old_stock, 'stock_after': new_stock})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/swap', methods=['POST'])
def swap_endpoint():
    """Simplified — Make.com only needs to tell us WHICH ROW changed in
    pending_approval. We figure out the customer/category/tier ourselves
    by reading that row directly, instead of asking Make.com to extract
    those pieces separately (which its Watch Changes trigger can't cleanly do)."""
    row_number = request.form.get('row_number')
    if not row_number:
        return jsonify({'error': 'row_number is required.'}), 400
    row_number = int(float(row_number))

    sheet_name = request.form.get('sheet_name', 'pending_approval')
    already_rejected = request.form.get('already_rejected', '')
    already_rejected = [s.strip() for s in already_rejected.split(',') if s.strip()]
    target_month = request.form.get('target_month') or None
    print(f'DEBUG /swap received: row_number={row_number}, sheet_name={sheet_name!r}, target_month={target_month!r}', flush=True)

    try:
        importlib.reload(recommend_v5)
        recommend_v5.set_target_month(target_month)

        ws = recommend_v5.wb[sheet_name]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        name_col = headers.index('Product Name') + 1
        cat_col = headers.index('Category') + 1
        tier_col = headers.index('Tier') + 1

        category = ws.cell(row_number, cat_col).value
        tier = ws.cell(row_number, tier_col).value
        current_product = ws.cell(row_number, name_col).value
        print(f'DEBUG /swap resolved: category={category!r}, tier={tier!r}, current_product={current_product!r}, workbook_path={recommend_v5.path!r}', flush=True)

        # The product currently sitting in this slot must never be
        # recommended back to itself as the "next alternative".
        if current_product and current_product not in already_rejected:
            already_rejected = already_rejected + [current_product]

        # Find which customer this row belongs to by scanning upward for
        # the nearest customer-header row above it.
        customer_name = None
        for r in range(row_number, 0, -1):
            fc = ws.cell(r, 1).value
            if isinstance(fc, str) and '|' in fc and 'Month:' in fc:
                customer_name = fc.split('|')[0].strip()
                break
        if not customer_name:
            return jsonify({'error': f'Could not find which customer owns row {row_number}.'}), 400

        # Also collect the other 4 products already in this customer's box,
        # so we don't recommend the same brand twice.
        current_box_products = []
        r = row_number
        while True:
            n = ws.cell(r, headers.index('#') + 1).value
            if not isinstance(n, (int, float)):
                break
            prod = ws.cell(r, name_col).value
            if r != row_number and prod:
                current_box_products.append(prod)
            r -= 1
        r = row_number + 1
        while True:
            n = ws.cell(r, headers.index('#') + 1).value
            if not isinstance(n, (int, float)):
                break
            prod = ws.cell(r, name_col).value
            if prod:
                current_box_products.append(prod)
            r += 1

        product, message = recommend_v5.get_next_alternative(
            customer_name, category, tier, already_rejected, current_box_products
        )
        if product is None:
            return jsonify({'found': False, 'message': message})

        return jsonify({
            'found': True,
            'product': {
                'name': product['name'],
                'category': product['category'],
                'tier': product['tier'],
                'stock': product['stock'],
                'price_aed': product.get('retail_price_aed'),
            },
            'message': message,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/generate_month', methods=['POST'])
def generate_month_endpoint():
    """Generates a full month's recommendations for EVERY active/due customer
    in ONE call — not one customer at a time. This matters: it shares stock
    tracking across all customers in the same call, so two different
    customers can never accidentally get recommended the same last-1-in-stock
    item within the same month's batch. Uses the bundled file directly —
    no file upload needed, avoiding the corruption risk of the Google Drive
    export method."""
    target_month = request.form.get('target_month')
    if not target_month:
        return jsonify({'error': 'target_month is required, e.g. 2026-09'}), 400

    try:
        importlib.reload(recommend_v5)
        recommend_v5.set_target_month(target_month)
        recommend_v5.reset_allocations()

        customers = recommend_v5.load_customers()
        all_rec, recent_boxes, box_timeline = recommend_v5.load_box_history()

        results = []
        for cid, cust in customers.items():
            due, reason = recommend_v5.is_customer_due(cust, box_timeline, target_month)
            if not due:
                continue
            out = recommend_v5.recommend(cust['name'])
            if isinstance(out, str):
                continue
            (c, picks, warnings, received, ratio, recent, hard_block, soft_avoid,
             hist_pat, total, timeline, inv_cat_map, value_summary) = out
            results.append({
                'customer_name': c['name'],
                'customer_id': c['id'],
                'box_type': c['box_type'],
                'products': [
                    {'name': p['name'], 'category': p['category'], 'tier': p['tier'],
                     'stock': p['stock'], 'price_aed': p.get('retail_price_aed')}
                    for p in picks
                ],
                'value_summary': value_summary,
                'warnings': warnings,
            })

        return jsonify({'target_month': target_month, 'customers': results})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/generate_month_rows', methods=['POST'])
def generate_month_rows_endpoint():
    """Returns a pre-flattened, ready-to-write list of rows — one header
    row per customer, immediately followed by their 5 product rows —
    already in the exact order they need to appear in the sheet. This
    means Make.com only needs ONE simple loop (the same reliable pattern
    already used tonight), not nested loops, avoiding real fragility."""
    target_month = request.form.get('target_month')
    if not target_month:
        return jsonify({'error': 'target_month is required, e.g. 2026-10'}), 400

    try:
        importlib.reload(recommend_v5)
        recommend_v5.set_target_month(target_month)
        recommend_v5.reset_allocations()

        customers = recommend_v5.load_customers()
        all_rec, recent_boxes, box_timeline = recommend_v5.load_box_history()

        rows = []
        for cid, cust in customers.items():
            due, reason = recommend_v5.is_customer_due(cust, box_timeline, target_month)
            if not due:
                continue
            out = recommend_v5.recommend(cust['name'])
            if isinstance(out, str):
                continue
            (c, picks, warnings, received, ratio, recent, hard_block, soft_avoid,
             hist_pat, total, timeline, inv_cat_map, value_summary) = out

            timeline_for_cust = box_timeline.get(cid, [])
            last_box = timeline_for_cust[-1][0] if timeline_for_cust else 'N/A'
            age = ''
            bday = cust.get('birthday')
            if bday:
                try:
                    from datetime import datetime as _dt
                    bdate = _dt.strptime(str(bday)[:10], '%Y-%m-%d')
                    today = _dt.now()
                    age = today.year - bdate.year - ((today.month, today.day) < (bdate.month, bdate.day))
                except Exception:
                    age = ''

            prefs = recommend_v5.load_preferences().get(cid, {})
            freqs = recommend_v5.load_frequencies().get(cid, {})
            quiz = recommend_v5.load_quiz().get(cid, {})

            header_text = f"  {c['name']}   |   {c['box_type']}   |   Last box: {last_box}   |   Month: {target_month}, Age {age}"

            rows.append({
                'row_type': 'header',
                'column_a': header_text,
                'customer_name': c['name'],
                'month': target_month,
            })
            for p in picks:
                why_text = recommend_v5.build_explanation(
                    p, cid, hard_block, soft_avoid, hist_pat, total,
                    timeline, inv_cat_map, prefs, freqs, quiz, age, cust.get('notes', '')
                )
                rows.append({
                    'row_type': 'product',
                    'product_name': p['name'],
                    'category': p['category'],
                    'tier': p['tier'],
                    'stock': p['stock'],
                    'price_aed': p.get('retail_price_aed'),
                    'why_recommended': why_text,
                    'customer_name': c['name'],
                    'month': target_month,
                })

        return jsonify({'target_month': target_month, 'rows': rows, 'row_count': len(rows)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

    target_month = request.form.get('target_month')
    if not target_month:
        return jsonify({'error': 'target_month is required, e.g. 2026-09'}), 400

    try:
        importlib.reload(recommend_v5)
        recommend_v5.set_target_month(target_month)
        recommend_v5.reset_allocations()

        customers = recommend_v5.load_customers()
        all_rec, recent_boxes, box_timeline = recommend_v5.load_box_history()

        results = []
        for cid, cust in customers.items():
            due, reason = recommend_v5.is_customer_due(cust, box_timeline, target_month)
            if not due:
                continue
            out = recommend_v5.recommend(cust['name'])
            if isinstance(out, str):
                continue
            (c, picks, warnings, received, ratio, recent, hard_block, soft_avoid,
             hist_pat, total, timeline, inv_cat_map, value_summary) = out
            results.append({
                'customer_name': c['name'],
                'customer_id': c['id'],
                'box_type': c['box_type'],
                'products': [
                    {'name': p['name'], 'category': p['category'], 'tier': p['tier'],
                     'stock': p['stock'], 'price_aed': p.get('retail_price_aed')}
                    for p in picks
                ],
                'value_summary': value_summary,
                'warnings': warnings,
            })

        return jsonify({'target_month': target_month, 'customers': results})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/generate_month_formatted', methods=['POST'])
def generate_month_formatted_endpoint():
    """Does the whole job in one call: generates recommendations for every
    due customer AND writes them into a properly formatted sheet (pink
    headers, purple customer rows — same style as every month built so
    far), reusing the same tested build_month_sheet() function. Returns
    the ready-to-use file — Make.com just uploads it back to replace the
    live Google Sheet, no Iterator or row-by-row writing needed at all."""
    if 'file' not in request.files:
        return jsonify({'error': 'No spreadsheet file was sent.'}), 400
    target_month = request.form.get('target_month')
    sheet_name = request.form.get('sheet_name')
    if not target_month:
        return jsonify({'error': 'target_month is required, e.g. 2026-10'}), 400
    if not sheet_name:
        sheet_name = f'pending_approval_{target_month.replace("-", "_")}'

    uploaded = request.files['file']
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        uploaded.save(tmp.name)
        tmp_path = tmp.name

    try:
        import build_final_sheet
        importlib.reload(recommend_v5)
        importlib.reload(build_final_sheet)
        recommend_v5.set_workbook_path(tmp_path)
        recommend_v5.set_target_month(target_month)
        build_final_sheet.set_sheet_path(tmp_path)

        build_final_sheet.build_month_sheet(target_month, sheet_name)

        from flask import send_file
        return send_file(tmp_path, as_attachment=True,
                          download_name='snatched_beauty_box_master_updated.xlsx')
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/finalize_item', methods=['POST'])
def finalize_item_endpoint():
    """Called once per approved row, using data Make.com already has from
    a simple Search Rows step — no file export needed at all, avoiding
    the binary/Base64 corruption risk entirely. Since Handle Approval
    already decremented stock in real time as things were approved, this
    only needs to create the permanent box/box_item record — not touch
    stock again."""
    customer_name = request.form.get('customer_name')
    customer_id = request.form.get('customer_id')
    box_type = request.form.get('box_type')
    product_name = request.form.get('product_name')
    target_month = request.form.get('target_month')
    box_id = request.form.get('box_id')  # same for all 5 items in one customer's box

    if not all([customer_name, customer_id, product_name, target_month, box_id]):
        return jsonify({'error': 'customer_name, customer_id, product_name, target_month, and box_id are all required.'}), 400

    try:
        importlib.reload(recommend_v5)
        ws_items = recommend_v5.wb['box_items']
        next_item_num = ws_items.max_row
        next_item_num += 1
        ws_items.append([f'BI{next_item_num:05d}', customer_name, customer_id, target_month, box_type, box_id, product_name])

        recommend_v5.wb.save(recommend_v5.path)
        push_to_github_async()
        return jsonify({'success': True, 'item_added': product_name, 'box_id': box_id})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/create_box_record', methods=['POST'])
def create_box_record_endpoint():
    """Called once per customer (not per item) to create the parent 'box'
    record. Returns a real box_id for Make.com to reuse across that
    customer's 5 separate /finalize_item calls."""
    customer_name = request.form.get('customer_name')
    customer_id = request.form.get('customer_id')
    box_type = request.form.get('box_type')
    target_month = request.form.get('target_month')

    if not all([customer_name, customer_id, target_month]):
        return jsonify({'error': 'customer_name, customer_id, and target_month are all required.'}), 400

    try:
        importlib.reload(recommend_v5)
        ws_boxes = recommend_v5.wb['boxes']
        next_box_num = ws_boxes.max_row + 1
        box_id = f'BOX{next_box_num:04d}'
        ws_boxes.append([box_id, customer_id, customer_name, target_month, 'sent', box_type])

        recommend_v5.wb.save(recommend_v5.path)
        push_to_github_async()
        return jsonify({'success': True, 'box_id': box_id})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# =========================================================================
# DASHBOARD — everything below powers the real dashboard (not Make.com).
# It all reads/writes ONE working sheet ("dashboard_working") that lives
# inside the same bundled workbook everything else already uses — no
# Google Sheets involved, no file upload/export round-trip. Every action
# here is self-contained: it writes its own result straight into the sheet,
# so state survives a page reload, a phone-to-laptop switch, or a restart.
# =========================================================================

DASH_SHEET = 'dashboard_working'
DASH_HEADERS = ['#', 'Product Name', 'Category', 'Tier', 'Stock', 'Why', 'Status', 'Manual', 'Price']

def _max_id_suffix(ws, col, prefix):
    """Returns the highest numeric suffix currently used by IDs like
    'BOX0188' / 'BI00940' in the given column — robust against workbooks
    that have blank formatted rows extending past the real last data row
    (ws.max_row can't be trusted in that case)."""
    best = 0
    for row in ws.iter_rows(min_row=2, min_col=col, max_col=col, values_only=True):
        val = row[0]
        if isinstance(val, str) and val.startswith(prefix):
            try:
                best = max(best, int(val[len(prefix):]))
            except ValueError:
                continue
    return best

def _get_dash_sheet(create=False):
    wb = recommend_v5.wb
    if DASH_SHEET not in wb.sheetnames:
        if not create:
            return None
        ws = wb.create_sheet(DASH_SHEET)
        ws.append(DASH_HEADERS)
        return ws
    return wb[DASH_SHEET]

def _parse_dash_sheet(ws):
    """Reads the whole working sheet back into structured JSON. Header rows
    store customer_id/box_type/month in columns B/C/D (not shown to a
    human — this sheet is internal-only); item rows store the actual
    product fields. Returns a list of customer blocks in sheet order."""
    if ws is None:
        return []
    customers = []
    current = None
    for r in range(2, ws.max_row + 1):
        col_a = ws.cell(r, 1).value
        if col_a is None or col_a == '':
            continue
        if isinstance(col_a, str) and col_a.startswith('HEADER:'):
            if current:
                customers.append(current)
            current = {
                'customer_name': col_a[len('HEADER:'):].strip(),
                'customer_id': ws.cell(r, 2).value,
                'box_type': ws.cell(r, 3).value,
                'month': ws.cell(r, 4).value,
                'header_row': r,
                'items': [],
            }
        else:
            # item row — column A holds the item index (1-5)
            if current is None:
                continue
            current['items'].append({
                'row_number': r,
                'idx': col_a,
                'name': ws.cell(r, 2).value,
                'category': ws.cell(r, 3).value,
                'tier': ws.cell(r, 4).value,
                'stock': ws.cell(r, 5).value,
                'why': ws.cell(r, 6).value,
                'status': ws.cell(r, 7).value or 'Pending',
                'manual': bool(ws.cell(r, 8).value),
                'price': ws.cell(r, 9).value,
            })
    if current:
        customers.append(current)
    return customers

def _find_item_row_context(ws, row_number):
    """Given a row number for one item, returns (customer_block, item_dict)
    so actions can see the customer_id/other-items-in-this-box context
    needed for brand-duplicate checks, without re-parsing the whole sheet
    by hand each time."""
    blocks = _parse_dash_sheet(ws)
    for b in blocks:
        for it in b['items']:
            if it['row_number'] == row_number:
                return b, it
    return None, None

@app.route('/dashboard/generate_month', methods=['POST'])
@dashboard_login_required
def dashboard_generate_month():
    """Builds fresh recommendations for every due customer and writes them
    into the working sheet, replacing whatever was there before. Refuses
    to clobber a month that still has unapproved work unless force=true —
    generating twice by accident shouldn't silently throw away swaps."""
    target_month = request.form.get('target_month')
    force = request.form.get('force') == 'true'
    if not target_month:
        return jsonify({'error': 'target_month is required, e.g. 2026-11'}), 400

    try:
        importlib.reload(recommend_v5)
        wb = recommend_v5.wb

        existing = _get_dash_sheet(create=False)
        existing_blocks = _parse_dash_sheet(existing) if (existing is not None and existing.max_row > 1) else []

        if existing_blocks and not force:
            still_open = [b['customer_name'] for b in existing_blocks if b.get('month') != target_month
                          or any(i['status'] != 'Approved' for i in b['items'])]
            if still_open:
                return jsonify({
                    'error': 'There is unfinished work already on the board.',
                    'unfinished_customers': still_open,
                    'hint': 'Pass force=true to overwrite it anyway.',
                }), 409

        # Any item still sitting on the board as 'Approved' has already had
        # its stock decremented (see /dashboard/approve). If we're about to
        # wipe the board — either because everything was Approved, or
        # because force=true is throwing away in-progress work — that stock
        # must go back into inventory first, or it's gone for good with
        # nothing to show for it.
        if existing_blocks:
            approved_names = [i['name'] for b in existing_blocks for i in b['items'] if i['status'] == 'Approved']
            if approved_names:
                ws_inv = wb['inventory']
                hi = [ws_inv.cell(1, c).value for c in range(1, ws_inv.max_column + 1)]
                n_col = hi.index('name') + 1
                s_col = hi.index('stock_qty') + 1
                inv_rows_by_name = {}
                for row in ws_inv.iter_rows(min_row=2):
                    nm = row[n_col - 1].value
                    if nm:
                        inv_rows_by_name[nm] = row
                for name in approved_names:
                    row = inv_rows_by_name.get(name)
                    if row:
                        row[s_col - 1].value = (row[s_col - 1].value or 0) + 1

        if DASH_SHEET in wb.sheetnames:
            del wb[DASH_SHEET]
        ws = _get_dash_sheet(create=True)

        recommend_v5.set_target_month(target_month)
        recommend_v5.reset_allocations()
        customers = recommend_v5.load_customers()
        all_rec, recent_boxes, box_timeline = recommend_v5.load_box_history()

        # Loaded ONCE here instead of once per customer inside the loop below.
        # None of this changes while the loop runs (nothing in recommend()
        # writes back to the sheet), so re-reading it fresh for every single
        # customer was pure wasted time -- with ~15-20 real customers this
        # was slow enough to blow past Render's 30s worker timeout and crash
        # the whole request (that's the SystemExit/SIGKILL you saw in the
        # logs). Reusing one snapshot for everyone fixes that at the root
        # instead of just asking Render to wait longer.
        preloaded = {
            'customers': customers,
            'box_history': (all_rec, recent_boxes, box_timeline),
            'prefs': recommend_v5.load_preferences(),
            'freqs': recommend_v5.load_frequencies(),
            'quiz': recommend_v5.load_quiz(),
            'inventory': recommend_v5.load_inventory(),
            'rejected': recommend_v5.load_rejected_products(),
            'blocked_brands': recommend_v5.load_blocked_brands(),
        }

        built = 0
        skipped = []
        for cid, cust in customers.items():
            due, reason = recommend_v5.is_customer_due(cust, box_timeline, target_month)
            if not due:
                skipped.append({'name': cust['name'], 'reason': reason})
                continue
            out = recommend_v5.recommend(cust['name'], _preloaded=preloaded)
            if isinstance(out, str):
                # recommend() returns a plain string instead of raising when
                # it can't build a box for a real, known reason (e.g. an
                # unset/unrecognized box type) -- most commonly a new or
                # tier-unconfirmed customer like "add Darya once she
                # confirms her tier." Surface it instead of failing silently.
                skipped.append({'name': cust['name'], 'reason': out})
                continue
            (c, picks, warnings, received, ratio, recent, hard_block, soft_avoid,
             hist_pat, total, timeline, inv_cat_map, value_summary) = out

            prefs = preloaded['prefs'].get(cid, {})
            freqs = preloaded['freqs'].get(cid, {})
            quiz = preloaded['quiz'].get(cid, {})
            age = ''
            bday = cust.get('birthday')
            if bday:
                try:
                    from datetime import datetime as _dt
                    bdate = _dt.strptime(str(bday)[:10], '%Y-%m-%d')
                    today = _dt.now()
                    age = today.year - bdate.year - ((today.month, today.day) < (bdate.month, bdate.day))
                except Exception:
                    age = ''

            ws.append([f'HEADER: {c["name"]}', cid, c['box_type'], target_month])
            for i, p in enumerate(picks, start=1):
                why_text = recommend_v5.build_explanation(
                    p, cid, hard_block, soft_avoid, hist_pat, total,
                    timeline, inv_cat_map, prefs, freqs, quiz, age, cust.get('notes', '')
                )
                ws.append([i, p['name'], p['category'], p['tier'], p['stock'],
                           why_text, 'Pending', False, p.get('retail_price_aed')])
            built += 1

        wb.save(recommend_v5.path)
        push_to_github_async()
        return jsonify({
            'success': True, 'target_month': target_month, 'customers_built': built,
            'skipped': skipped,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/dashboard/generate_for_customer', methods=['POST'])
@dashboard_login_required
def dashboard_generate_for_customer():
    """Builds a box for exactly ONE customer and adds her to the working
    board, without touching anyone else already there or re-building
    anyone whose box for this month is already committed. For the exact
    case of a customer who was skipped from the normal month generation
    (e.g. her tier wasn't confirmed yet) -- once she confirms, use this
    to add just her, any time, even after the rest of the month has
    already been committed."""
    customer_id = request.form.get('customer_id')
    target_month = request.form.get('target_month')
    if not customer_id or not target_month:
        return jsonify({'error': 'customer_id and target_month are required.'}), 400

    try:
        importlib.reload(recommend_v5)
        wb = recommend_v5.wb

        customers = recommend_v5.load_customers()
        cust = customers.get(customer_id)
        if not cust:
            return jsonify({'error': f'No customer found with id "{customer_id}".'}), 404

        # Refuse if she already has a real, committed box for this exact
        # month -- the #1 thing this endpoint exists to prevent.
        ws_boxes = wb['boxes']
        for row in ws_boxes.iter_rows(min_row=2, values_only=True):
            if row[1] == customer_id and str(row[3]) == target_month and str(row[4]).strip().lower() != 'cancelled':
                return jsonify({'error': f'{cust["name"]} already has a committed box for {target_month} ({row[0]}).'}), 409

        existing = _get_dash_sheet(create=False)
        existing_blocks = _parse_dash_sheet(existing) if existing is not None else []
        if any(b['customer_id'] == customer_id and b.get('month') == target_month for b in existing_blocks):
            return jsonify({'error': f'{cust["name"]} is already on the board for {target_month}.'}), 409

        recommend_v5.set_target_month(target_month)
        recommend_v5.reset_allocations()
        # reset_allocations() clears the per-run duplicate-brand/subcategory
        # tracking that's meant to keep ONE MONTH's boxes from repeating the
        # same brand across customers -- since this runs standalone (not as
        # part of a full-month build), it starts with a clean slate rather
        # than sharing state with whatever was last generated.

        out = recommend_v5.recommend(cust['name'])
        if isinstance(out, str):
            return jsonify({'error': out}), 400
        (c, picks, warnings, received, ratio, recent, hard_block, soft_avoid,
         hist_pat, total, timeline, inv_cat_map, value_summary) = out

        prefs = recommend_v5.load_preferences().get(customer_id, {})
        freqs = recommend_v5.load_frequencies().get(customer_id, {})
        quiz = recommend_v5.load_quiz().get(customer_id, {})
        age = ''
        bday = cust.get('birthday')
        if bday:
            try:
                bdate = _datetime.strptime(str(bday)[:10], '%Y-%m-%d')
                today = _datetime.now()
                age = today.year - bdate.year - ((today.month, today.day) < (bdate.month, bdate.day))
            except Exception:
                age = ''

        ws = _get_dash_sheet(create=True)
        ws.append([f'HEADER: {c["name"]}', customer_id, c['box_type'], target_month])
        for i, p in enumerate(picks, start=1):
            why_text = recommend_v5.build_explanation(
                p, customer_id, hard_block, soft_avoid, hist_pat, total,
                timeline, inv_cat_map, prefs, freqs, quiz, age, cust.get('notes', '')
            )
            ws.append([i, p['name'], p['category'], p['tier'], p['stock'],
                       why_text, 'Pending', False, p.get('retail_price_aed')])

        wb.save(recommend_v5.path)
        push_to_github_async()
        return jsonify({'success': True, 'customer_name': c['name'], 'target_month': target_month})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/dashboard/state', methods=['GET'])
@dashboard_login_required
def dashboard_state():
    """Returns the full current board — every customer, every item, current
    status — for the review page to render. Called on page load and after
    the drawer/manual-entry flows need a fresh read."""
    try:
        importlib.reload(recommend_v5)
        ws = _get_dash_sheet(create=False)
        blocks = _parse_dash_sheet(ws)
        return jsonify({'customers': blocks})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

LOW_STOCK_THRESHOLD = 3  # a category with this many or fewer total units left,
                          # across every product in it, is flagged "critically low"

@app.route('/dashboard/home_stats', methods=['GET'])
@dashboard_login_required
def dashboard_home_stats():
    """Powers the two Home-page stat cards from the original mockup that
    weren't wired up yet: birthdays falling in the month currently being
    reviewed (matched on month-of-year, so it's correct every year — not
    tied to the calendar month you happen to log in), and inventory
    categories running critically low. Both use real data, no samples."""
    target_month = request.args.get('target_month')  # 'YYYY-MM', optional

    try:
        importlib.reload(recommend_v5)
        customers = recommend_v5.load_customers()

        birthdays = []
        if target_month:
            try:
                target_mm = int(target_month.split('-')[1])
            except (ValueError, IndexError):
                target_mm = None
            if target_mm:
                for c in customers.values():
                    bday = c.get('birthday')
                    if not bday:
                        continue
                    try:
                        bmonth = int(str(bday)[5:7])
                    except (ValueError, IndexError):
                        continue
                    if bmonth == target_mm and str(c.get('status', '')).strip().lower() == 'active':
                        birthdays.append({
                            'name': c['name'],
                            'birthday': str(bday)[:10],
                            'subscribed_since': c.get('subscribed_since'),
                        })
        birthdays.sort(key=lambda b: b['birthday'][8:10] if len(b['birthday']) >= 10 else '')

        inv = recommend_v5.load_inventory()
        totals = {}
        for p in inv:
            cat = p.get('category')
            if not cat:
                continue
            totals[cat] = totals.get(cat, 0) + (p.get('stock') or 0)
        low_stock = sorted(
            [{'category': cat, 'total_stock': qty} for cat, qty in totals.items() if qty <= LOW_STOCK_THRESHOLD],
            key=lambda x: x['total_stock']
        )

        return jsonify({'birthdays': birthdays, 'low_stock_categories': low_stock, 'low_stock_threshold': LOW_STOCK_THRESHOLD})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/dashboard/months', methods=['GET'])
@dashboard_login_required
def dashboard_months():
    """Every month that's ever had a box committed, newest first --
    powers the 'Past months' browser so Iqra can look back at any
    month's full history without hunting it one customer at a time."""
    try:
        importlib.reload(recommend_v5)
        ws_b = recommend_v5.wb['boxes']
        months = sorted({str(row[3]) for row in ws_b.iter_rows(min_row=2, values_only=True) if row[3]}, reverse=True)
        return jsonify({'months': months})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/dashboard/month_detail', methods=['GET'])
@dashboard_login_required
def dashboard_month_detail():
    """Full detail for one past (committed) month: every box that went
    out, who got it, and exactly what was in it -- everything Iqra
    would otherwise have to hunt for customer-by-customer. Cancelled
    boxes for that month are included too, but counted separately, so
    a month never silently looks like it shipped more than it did."""
    month = request.args.get('month')
    if not month:
        return jsonify({'error': 'month is required, e.g. ?month=2026-09'}), 400
    try:
        importlib.reload(recommend_v5)
        wb = recommend_v5.wb
        ws_b = wb['boxes']
        ws_i = wb['box_items']

        boxes = [
            {'box_id': row[0], 'customer_id': row[1], 'customer_name': row[2],
             'status': row[4], 'box_type': row[5]}
            for row in ws_b.iter_rows(min_row=2, values_only=True) if str(row[3]) == month
        ]
        if not boxes:
            return jsonify({'month': month, 'box_count': 0, 'sent_count': 0, 'cancelled_count': 0, 'boxes': []})

        items_by_box = {}
        for row in ws_i.iter_rows(min_row=2, values_only=True):
            if row[3] and str(row[3]) == month and row[5] and row[6]:
                items_by_box.setdefault(row[5], []).append(row[6])

        for b in boxes:
            b['products'] = items_by_box.get(b['box_id'], [])
        boxes.sort(key=lambda b: (b['customer_name'] or '').lower())

        sent_count = sum(1 for b in boxes if str(b['status']).strip().lower() != 'cancelled')
        cancelled_count = len(boxes) - sent_count

        return jsonify({
            'month': month, 'box_count': len(boxes), 'sent_count': sent_count,
            'cancelled_count': cancelled_count, 'boxes': boxes,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/dashboard/all_customers', methods=['GET'])
@dashboard_login_required
def dashboard_all_customers():
    """Every customer on file, regardless of whether a month's board is
    currently loaded — powers the Customers-tab search autocomplete so
    Iqra can find and cancel anyone any day, not just customers who
    happen to be on the most recently generated month's board."""
    try:
        importlib.reload(recommend_v5)
        customers = recommend_v5.load_customers()
        out = sorted(
            [{'id': c['id'], 'name': c['name'], 'status': c.get('status')} for c in customers.values()],
            key=lambda c: (c['name'] or '').lower()
        )
        return jsonify({'customers': out})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/dashboard/inventory', methods=['GET'])
@dashboard_login_required
def dashboard_inventory():
    """Powers two different things depending on the 'all' flag:
    - Default (no flag): only items with stock > 0 -- this is what the
      'Enter my own' manual-pick search box uses, same rule the manual-pick
      endpoint itself enforces (can't hand-pick something that's out of stock).
    - all=1: every product regardless of stock, including 0 -- this is what
      the 'Current inventory' browse/check panel uses, so Iqra can see
      out-of-stock items too and confirm counts after a restock."""
    try:
        importlib.reload(recommend_v5)
        inv = recommend_v5.load_inventory()
        show_all = request.args.get('all') == '1'
        items = [
            {'name': p['name'], 'category': p['category'], 'tier': p['tier'],
             'stock': p['stock'], 'price': p.get('retail_price_aed')}
            for p in inv if show_all or (p['stock'] or 0) > 0
        ]
        items.sort(key=lambda x: (x['name'] or '').lower())
        return jsonify({'items': items, 'count': len(items)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

def _apply_inventory_change(wb, ws, name, qty, category='', tier='', price_raw=None):
    """Core add-or-restock logic, shared by the single-item and bulk-upload
    endpoints. If `name` already exists in the inventory sheet (exact match,
    case-insensitive), its stock_qty is just increased by `qty` -- this is
    the normal 'restocking' case. If it's a brand-new product, a new row is
    created with a freshly-generated product_id, continuing the same
    NEW#### numbering already used for hand-added products in this sheet.
    Returns a result dict; does NOT save the workbook or push to GitHub --
    callers do that once, after all rows for a batch are applied, so a
    100-item bulk upload doesn't save/push 100 separate times."""
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    n_col = headers.index('name') + 1
    s_col = headers.index('stock_qty') + 1
    id_col = headers.index('product_id') + 1
    cat_col = headers.index('category') + 1
    tier_col = headers.index('tier') + 1
    price_col = headers.index('retail_price_aed') + 1

    name_lower = name.lower()
    max_num = 0
    for r in range(2, ws.max_row + 1):
        pid = ws.cell(r, id_col).value
        existing_name = ws.cell(r, n_col).value
        if pid:
            digits = ''.join(ch for ch in str(pid) if ch.isdigit())
            if digits:
                max_num = max(max_num, int(digits))
        if existing_name and str(existing_name).strip().lower() == name_lower:
            old_stock = ws.cell(r, s_col).value or 0
            new_stock = old_stock + qty
            ws.cell(r, s_col).value = new_stock
            return {
                'success': True, 'created_new': False, 'name': existing_name,
                'old_stock': old_stock, 'new_stock': new_stock,
            }

    new_id = f'NEW{max_num + 1:04d}'
    new_row_num = _next_append_row(ws, check_cols=(id_col, n_col))
    ws.cell(new_row_num, id_col).value = new_id
    ws.cell(new_row_num, n_col).value = name
    ws.cell(new_row_num, tier_col).value = tier
    ws.cell(new_row_num, cat_col).value = category
    ws.cell(new_row_num, s_col).value = qty
    if price_raw:
        try:
            ws.cell(new_row_num, price_col).value = float(price_raw)
        except ValueError:
            pass

    return {
        'success': True, 'created_new': True, 'product_id': new_id,
        'name': name, 'old_stock': 0, 'new_stock': qty,
    }

@app.route('/dashboard/add_inventory', methods=['POST'])
@dashboard_login_required
def dashboard_add_inventory():
    """Lets Iqra add new stock straight from the dashboard -- no
    spreadsheet needed. If the product name already exists (exact match,
    case-insensitive), its stock_qty is just increased by the amount
    given -- this is how restocking an item she already carries works,
    same as adding a brand-new one. See _apply_inventory_change()."""
    name = (request.form.get('name') or '').strip()
    qty_raw = request.form.get('quantity')
    category = (request.form.get('category') or '').strip()
    tier = (request.form.get('tier') or '').strip()
    price_raw = request.form.get('retail_price_aed')

    if not name:
        return jsonify({'error': 'name is required.'}), 400
    try:
        qty = int(qty_raw)
    except (TypeError, ValueError):
        return jsonify({'error': 'quantity must be a whole number.'}), 400
    if qty <= 0:
        return jsonify({'error': 'quantity must be greater than 0.'}), 400

    try:
        importlib.reload(recommend_v5)
        wb = recommend_v5.wb
        ws = wb['inventory']
        result = _apply_inventory_change(wb, ws, name, qty, category, tier, price_raw)
        wb.save(recommend_v5.path)
        push_to_github_async()
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/dashboard/bulk_add_inventory', methods=['POST'])
@dashboard_login_required
def dashboard_bulk_add_inventory():
    """Lets Iqra paste in a whole restock list at once instead of adding
    100 items one at a time. Expects a 'lines' form field, one product per
    line, formatted as:  Name, Quantity[, Category, Tier]
    (Category/Tier only matter for brand-new products; existing products
    just get restocked by name match.) Blank lines and lines starting with
    # are ignored. Every valid line is applied to the same in-memory
    workbook, then it's saved and pushed to GitHub exactly once."""
    raw = request.form.get('lines') or ''
    raw_lines = [ln.strip() for ln in raw.splitlines()]

    try:
        importlib.reload(recommend_v5)
        wb = recommend_v5.wb
        ws = wb['inventory']

        created, restocked, errors = [], [], []
        any_applied = False
        for i, line in enumerate(raw_lines, start=1):
            if not line or line.startswith('#'):
                continue
            parts = [p.strip() for p in line.split(',')]
            if len(parts) < 2:
                errors.append({'line': i, 'text': line, 'reason': 'Expected "Name, Quantity" — no comma found.'})
                continue
            name = parts[0]
            qty_raw = parts[1]
            category = parts[2] if len(parts) > 2 else ''
            tier = parts[3] if len(parts) > 3 else ''
            if not name:
                errors.append({'line': i, 'text': line, 'reason': 'Missing product name.'})
                continue
            try:
                qty = int(qty_raw)
            except ValueError:
                errors.append({'line': i, 'text': line, 'reason': f'"{qty_raw}" is not a whole number.'})
                continue
            if qty <= 0:
                errors.append({'line': i, 'text': line, 'reason': 'Quantity must be greater than 0.'})
                continue

            result = _apply_inventory_change(wb, ws, name, qty, category, tier)
            any_applied = True
            if result['created_new']:
                created.append(result)
            else:
                restocked.append(result)

        if any_applied:
            wb.save(recommend_v5.path)
            push_to_github_async()

        return jsonify({
            'success': True,
            'created': created,
            'restocked': restocked,
            'errors': errors,
            'created_count': len(created),
            'restocked_count': len(restocked),
            'error_count': len(errors),
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

def _next_append_row(ws, check_cols=(1, 2)):
    """Returns the row number to write a brand-new row into. Several
    sheets in this workbook have thousands of pre-formatted-but-empty
    trailing rows, which makes ws.max_row unreliable for finding 'the
    real last row of data' -- using it directly would leave a huge gap
    of blank rows before a freshly appended one. This scans for the
    actual last row that has something in it instead."""
    last = 1
    for r in range(2, ws.max_row + 1):
        if any(ws.cell(r, c).value not in (None, '') for c in check_cols):
            last = r
    return last + 1

def _restock_product_by_name(wb, product_name):
    """Puts one unit of a product back into inventory by name. Used
    whenever a row that was already Approved (and had therefore already
    had its stock taken, per /dashboard/approve) is about to be
    overwritten with a different product by a swap or manual pick — the
    old product's stock must come back, or it's gone for good even
    though she never actually kept it in the box."""
    if not product_name:
        return False
    ws_inv = wb['inventory']
    hi = [ws_inv.cell(1, c).value for c in range(1, ws_inv.max_column + 1)]
    n_col = hi.index('name') + 1
    s_col = hi.index('stock_qty') + 1
    for row in ws_inv.iter_rows(min_row=2):
        if row[n_col - 1].value == product_name:
            row[s_col - 1].value = (row[s_col - 1].value or 0) + 1
            return True
    return False

def _write_item_row(ws, row_number, product, why, manual=False):
    ws.cell(row_number, 2).value = product['name']
    ws.cell(row_number, 3).value = product['category']
    ws.cell(row_number, 4).value = product['tier']
    ws.cell(row_number, 5).value = product['stock']
    ws.cell(row_number, 6).value = why
    ws.cell(row_number, 7).value = 'Pending'  # any change re-opens it for approval
    ws.cell(row_number, 8).value = manual
    ws.cell(row_number, 9).value = product.get('retail_price_aed', product.get('price'))

@app.route('/dashboard/swap', methods=['POST'])
@dashboard_login_required
def dashboard_swap():
    """Same-category swap — the existing, already-tested alternative-finding
    logic, but self-contained: finds the next alternative AND writes it
    straight into the sheet in one call, instead of returning a suggestion
    for something else to write down separately."""
    row_number = request.form.get('row_number')
    if not row_number:
        return jsonify({'error': 'row_number is required.'}), 400
    row_number = int(row_number)

    try:
        importlib.reload(recommend_v5)
        ws = _get_dash_sheet(create=False)
        block, item = _find_item_row_context(ws, row_number)
        if not item:
            return jsonify({'error': f'No item found at row {row_number}.'}), 404
        recommend_v5.set_target_month(block['month'])  # eligibility math must use the
                                                         # month being reviewed, not today's

        other_products = [i['name'] for i in block['items'] if i['row_number'] != row_number]
        product, message = recommend_v5.get_next_alternative(
            block['customer_name'], item['category'], item['tier'],
            already_rejected=[item['name']], current_box_products=other_products
        )
        if product is None:
            return jsonify({'found': False, 'message': message})

        # If she'd already approved this slot (stock already taken for it,
        # per /dashboard/approve) and only now decided to swap it, the
        # product she's replacing must go back into inventory — otherwise
        # it's permanently gone even though it never actually shipped.
        if item['status'] == 'Approved':
            _restock_product_by_name(recommend_v5.wb, item['name'])

        _write_item_row(ws, row_number, product,
                         'Owner-requested swap — next best match for her preferences, timing, and brand rules.')
        recommend_v5.wb.save(recommend_v5.path)
        push_to_github_async()
        updated = next(i for i in _find_item_row_context(ws, row_number)[0]['items'] if i['row_number'] == row_number)
        return jsonify({'found': True, 'item': updated})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/dashboard/swap_category', methods=['POST'])
@dashboard_login_required
def dashboard_swap_category():
    """Different-category swap — same real eligibility engine (stock,
    history, frequency, brand rules), just pointed at a category she picked
    instead of the slot's original one. Keeps her plan's tier (Essentials/
    Prestige) fixed since that's billing-related, only category changes."""
    row_number = request.form.get('row_number')
    new_category = request.form.get('category')
    if not row_number or not new_category:
        return jsonify({'error': 'row_number and category are both required.'}), 400
    row_number = int(row_number)

    try:
        importlib.reload(recommend_v5)
        ws = _get_dash_sheet(create=False)
        block, item = _find_item_row_context(ws, row_number)
        if not item:
            return jsonify({'error': f'No item found at row {row_number}.'}), 404
        recommend_v5.set_target_month(block['month'])
        old_category = item['category']

        other_products = [i['name'] for i in block['items'] if i['row_number'] != row_number]
        product, message = recommend_v5.get_next_alternative(
            block['customer_name'], new_category, item['tier'],
            already_rejected=[], current_box_products=other_products
        )
        if product is None:
            return jsonify({'found': False, 'message': message})

        if item['status'] == 'Approved':
            _restock_product_by_name(recommend_v5.wb, item['name'])

        _write_item_row(ws, row_number, product,
                         f'Owner-requested category change — swapped from "{old_category}" to "{new_category}".')
        recommend_v5.wb.save(recommend_v5.path)
        push_to_github_async()
        updated = next(i for i in _find_item_row_context(ws, row_number)[0]['items'] if i['row_number'] == row_number)
        return jsonify({'found': True, 'item': updated, 'old_category': old_category, 'new_category': new_category})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/dashboard/eligible_products', methods=['GET'])
@dashboard_login_required
def dashboard_eligible_products():
    """Powers the 'Enter my own' picker list. Given a row_number, returns
    every real, in-stock product that matches the customer's tier (for a
    Mixed box, that means both Essentials and Prestige, since that's what
    her box actually draws from) -- NOT the whole catalog.

    Products from a brand she has strictly blocked are NOT removed from
    this list -- Iqra asked for them to still be visible but clearly
    flagged, so she can see at a glance what NOT to pick rather than
    wondering why something is missing. The automatic recommend/swap
    engine (recommend_v5.recommend, find_alternative) already hard-excludes
    blocked brands on its own and always has -- this list is a manual
    override tool, so it warns instead of hiding."""
    row_number = request.args.get('row_number')
    if not row_number:
        return jsonify({'error': 'row_number is required.'}), 400
    row_number = int(row_number)

    try:
        importlib.reload(recommend_v5)
        ws = _get_dash_sheet(create=False)
        block, item = _find_item_row_context(ws, row_number)
        if not item:
            return jsonify({'error': f'No item found at row {row_number}.'}), 404

        customers = recommend_v5.load_customers()
        cust = next((c for c in customers.values() if c['name'] == block['customer_name']), None)
        box_type = cust['box_type'] if cust else None
        ratio = recommend_v5.BOX_RATIOS.get(box_type)
        eligible_tiers = set(ratio.keys()) if ratio else ({box_type} if box_type else set())

        blocked_brands = recommend_v5.load_blocked_brands().get(cust['id'], set()) if cust else set()
        current_names_lower = {i['name'].strip().lower() for i in block['items']}

        inventory = recommend_v5.load_inventory()
        items = []
        for p in inventory:
            if (p['stock'] or 0) <= 0:
                continue
            if eligible_tiers and p['tier'] not in eligible_tiers:
                continue
            brand = recommend_v5.get_brand(p['name'])
            items.append({
                'name': p['name'], 'category': p['category'], 'tier': p['tier'],
                'stock': p['stock'], 'price': p.get('retail_price_aed'),
                'brand': brand,
                'blocked': brand in blocked_brands,
                'in_box': p['name'].strip().lower() in current_names_lower,
            })
        items.sort(key=lambda x: (x['category'] or '', x['name'] or ''))

        return jsonify({
            'items': items,
            'total': len(items),
            'blocked_count': sum(1 for i in items if i['blocked']),
            'box_type': box_type,
            'eligible_tiers': sorted(eligible_tiers),
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/dashboard/manual_pick', methods=['POST'])
@dashboard_login_required
def dashboard_manual_pick():
    """Owner types in a specific product herself — no scoring/eligibility
    engine involved, since she's overriding it on purpose. Still validates
    the product is real and in stock, and still runs the brand-duplicate
    and repeat-history safety checks so she sees a warning if either applies
    (she can still proceed either way — it's her call)."""
    row_number = request.form.get('row_number')
    product_name = request.form.get('product_name')
    if not row_number or not product_name:
        return jsonify({'error': 'row_number and product_name are both required.'}), 400
    row_number = int(row_number)

    try:
        importlib.reload(recommend_v5)
        ws = _get_dash_sheet(create=False)
        block, item = _find_item_row_context(ws, row_number)
        if not item:
            return jsonify({'error': f'No item found at row {row_number}.'}), 404

        inventory = recommend_v5.load_inventory()
        product = next((p for p in inventory if p['name'].lower() == product_name.lower()), None)
        if not product:
            return jsonify({'error': f'"{product_name}" is not a real, current inventory item.'}), 400
        if (product['stock'] or 0) < 1:
            return jsonify({'error': f'"{product_name}" is out of stock right now.'}), 400

        new_brand = recommend_v5.get_brand(product['name'])
        brand_clash = any(
            recommend_v5.get_brand(i['name']) == new_brand
            for i in block['items'] if i['row_number'] != row_number
        )

        customers = recommend_v5.load_customers()
        cust = next((c for c in customers.values() if c['name'] == block['customer_name']), None)
        all_rec, _, _ = recommend_v5.load_box_history()
        repeat = bool(cust) and product['name'].lower() in all_rec.get(cust['id'], set())
        blocked = bool(cust) and new_brand in recommend_v5.load_blocked_brands().get(cust['id'], set())

        if item['status'] == 'Approved' and item['name'] != product['name']:
            _restock_product_by_name(recommend_v5.wb, item['name'])

        _write_item_row(ws, row_number, product,
                         'Owner-selected manually — chosen directly instead of an engine suggestion.',
                         manual=True)
        recommend_v5.wb.save(recommend_v5.path)
        push_to_github_async()

        updated = next(i for i in _find_item_row_context(ws, row_number)[0]['items'] if i['row_number'] == row_number)
        return jsonify({'success': True, 'item': updated, 'brand_clash': brand_clash, 'repeat': repeat, 'blocked': blocked, 'brand': new_brand})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/dashboard/approve', methods=['POST'])
@dashboard_login_required
def dashboard_approve():
    """Marks one item approved and decrements its stock immediately —
    same real-time-stock-tracking behavior as the existing approve_item,
    just also setting the Status column so state is self-contained."""
    row_number = request.form.get('row_number')
    if not row_number:
        return jsonify({'error': 'row_number is required.'}), 400
    row_number = int(row_number)

    try:
        importlib.reload(recommend_v5)
        ws = _get_dash_sheet(create=False)
        block, item = _find_item_row_context(ws, row_number)
        if not item:
            return jsonify({'error': f'No item found at row {row_number}.'}), 404
        if item['status'] == 'Approved':
            return jsonify({'success': True, 'already_approved': True})

        ws_inv = recommend_v5.wb['inventory']
        hi = [ws_inv.cell(1, c).value for c in range(1, ws_inv.max_column + 1)]
        n_col = hi.index('name') + 1
        s_col = hi.index('stock_qty') + 1
        updated = False
        for row in ws_inv.iter_rows(min_row=2):
            if row[n_col - 1].value == item['name']:
                old_stock = row[s_col - 1].value or 0
                # Stock is only ever checked at PICK time (manual_pick, swap),
                # never re-checked here at approve time -- so two different
                # customers can both end up with the same last-1-in-stock
                # product sitting Pending at once (nothing reserves it when
                # picked, only Approve actually decrements). Without this
                # check, the second Approve would silently clamp to 0 and
                # mark her box ready-to-ship for a product that doesn't
                # physically exist anymore. Refuse instead, same as
                # manual_pick already refuses to pick something with 0 stock.
                if old_stock < 1:
                    return jsonify({
                        'error': f'"{item["name"]}" is out of stock right now (someone else likely got the last one approved first). '
                                 f'Swap or pick a different product for this slot before approving.',
                    }), 409
                row[s_col - 1].value = old_stock - 1
                updated = True
                break
        if not updated:
            return jsonify({'error': f'Product "{item["name"]}" not found in inventory.'}), 400

        ws.cell(row_number, 7).value = 'Approved'
        recommend_v5.wb.save(recommend_v5.path)
        push_to_github_async()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/dashboard/unapprove', methods=['POST'])
@dashboard_login_required
def dashboard_unapprove():
    """Reverses /dashboard/approve — restores the stock it took."""
    row_number = request.form.get('row_number')
    if not row_number:
        return jsonify({'error': 'row_number is required.'}), 400
    row_number = int(row_number)

    try:
        importlib.reload(recommend_v5)
        ws = _get_dash_sheet(create=False)
        block, item = _find_item_row_context(ws, row_number)
        if not item:
            return jsonify({'error': f'No item found at row {row_number}.'}), 404
        if item['status'] != 'Approved':
            return jsonify({'success': True, 'already_pending': True})

        ws_inv = recommend_v5.wb['inventory']
        hi = [ws_inv.cell(1, c).value for c in range(1, ws_inv.max_column + 1)]
        n_col = hi.index('name') + 1
        s_col = hi.index('stock_qty') + 1
        for row in ws_inv.iter_rows(min_row=2):
            if row[n_col - 1].value == item['name']:
                row[s_col - 1].value = (row[s_col - 1].value or 0) + 1
                break

        ws.cell(row_number, 7).value = 'Pending'
        recommend_v5.wb.save(recommend_v5.path)
        push_to_github_async()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/dashboard/download_workbook', methods=['GET'])
@dashboard_login_required
def dashboard_download_workbook():
    """Lets Iqra download the actual live spreadsheet, exactly as it is
    on the server right now — every customer, every box ever committed,
    full inventory — as one .xlsx file she can open in Excel or Google
    Sheets any time she wants the whole picture in one place instead of
    looking customer-by-customer. This is a snapshot at download time,
    not a live-synced sheet — nothing on her computer stays connected to
    the server after she downloads it; to see anything fresher she just
    downloads again."""
    try:
        importlib.reload(recommend_v5)
        from flask import send_file
        return send_file(
            recommend_v5.path, as_attachment=True,
            download_name=f'snatched_beauty_box_master_{_datetime.now().strftime("%Y-%m-%d")}.xlsx'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/dashboard/commit_month', methods=['POST'])
@dashboard_login_required
def dashboard_commit_month():
    """Commits every customer on the board who is fully approved (all 5
    items) — writes permanent boxes/box_items records (same schema
    /create_box_record and /finalize_item already use) and removes them
    from the working sheet so they can't be double-committed. Customers
    who aren't fully approved yet are left on the board untouched."""
    target_month = request.form.get('target_month')
    if not target_month:
        return jsonify({'error': 'target_month is required.'}), 400

    try:
        importlib.reload(recommend_v5)
        wb = recommend_v5.wb
        ws = _get_dash_sheet(create=False)
        blocks = _parse_dash_sheet(ws)

        ws_boxes = wb['boxes']
        ws_items = wb['box_items']
        # Scan actual IDs rather than trusting max_row — some exported copies
        # of this workbook carry blank formatted rows well past the real
        # last row, which would silently skip hundreds of row numbers.
        next_box_num = _max_id_suffix(ws_boxes, 1, 'BOX')
        next_item_num = _max_id_suffix(ws_items, 1, 'BI')

        committed = []
        skipped = []
        rows_to_delete = []

        for b in blocks:
            if b.get('month') != target_month:
                continue
            if not b['items'] or any(i['status'] != 'Approved' for i in b['items']):
                skipped.append(b['customer_name'])
                continue

            next_box_num += 1
            box_id = f'BOX{next_box_num:04d}'
            ws_boxes.append([box_id, b['customer_id'], b['customer_name'], target_month, 'sent', b['box_type']])
            for it in b['items']:
                next_item_num += 1
                ws_items.append([f'BI{next_item_num:05d}', b['customer_name'], b['customer_id'],
                                  target_month, b['box_type'], box_id, it['name']])

            committed.append(b['customer_name'])
            rows_to_delete.append(b['header_row'])
            rows_to_delete.extend(i['row_number'] for i in b['items'])

        for r in sorted(rows_to_delete, reverse=True):
            ws.delete_rows(r, 1)

        wb.save(recommend_v5.path)
        push_to_github_async()
        return jsonify({
            'success': True,
            'target_month': target_month,
            'committed_customers': committed,
            'still_pending_customers': skipped,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


def commit_month_endpoint():
    """Called when the owner clicks 'Approve & Lock In' for a month.
    Reads the CURRENT state of the spreadsheet directly — whatever she's
    already approved/swapped in the pending_approval sheet for that month —
    and commits exactly those products. Make.com's job is simple: export
    the current sheet, send it here, upload back whatever comes back.
    No complex data-building required on the Make.com side at all."""
    if 'file' not in request.files:
        return jsonify({'error': 'No spreadsheet file was sent.'}), 400
    target_month = request.form.get('target_month')
    sheet_name = request.form.get('sheet_name', 'pending_approval')
    if not target_month:
        return jsonify({'error': 'target_month is required.'}), 400

    uploaded = request.files['file']
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        uploaded.save(tmp.name)
        tmp_path = tmp.name

    try:
        import openpyxl as oxl
        from collections import Counter, defaultdict
        wb = oxl.load_workbook(tmp_path)

        ws_pa = wb[sheet_name]
        headers = [ws_pa.cell(1, c).value for c in range(1, ws_pa.max_column + 1)]
        num_col = headers.index('#') + 1
        name_col = headers.index('Product Name') + 1
        status_col = headers.index('STATUS ✏️') + 1

        current_customer = None
        approved = defaultdict(list)
        for row in ws_pa.iter_rows(min_row=1, values_only=True):
            fc = row[0]
            if isinstance(fc, str) and '|' in fc and 'Month:' in fc:
                current_customer = fc.split('|')[0].strip()
                continue
            n = row[num_col-1]
            if isinstance(n, (int, float)) and row[status_col-1] == 'approved':
                approved[current_customer].append(row[name_col-1])

        wsc = wb['customers']
        hc = [wsc.cell(1, c).value for c in range(1, wsc.max_column + 1)]
        name_col_c = hc.index('name') + 1
        cid_col = hc.index('customer_id') + 1
        box_col = hc.index('box_type') + 1
        name_to_id, name_to_boxtype = {}, {}
        for r in wsc.iter_rows(min_row=2, values_only=True):
            if r[name_col_c-1]:
                name_to_id[r[name_col_c-1]] = r[cid_col-1]
                name_to_boxtype[r[name_col_c-1]] = r[box_col-1]

        ws_boxes = wb['boxes']
        ws_items = wb['box_items']
        next_box_num = ws_boxes.max_row + 1
        next_item_num = ws_items.max_row

        usage = Counter()
        committed_customers = []
        for cust_name, products in approved.items():
            cid = name_to_id.get(cust_name)
            if not cid or not products:
                continue
            box_id = f'BOX{next_box_num:04d}'
            next_box_num += 1
            box_type = name_to_boxtype.get(cust_name)
            ws_boxes.append([box_id, cid, cust_name, target_month, 'sent', box_type])
            for prod in products:
                next_item_num += 1
                ws_items.append([f'BI{next_item_num:05d}', cust_name, cid, target_month, box_type, box_id, prod])
                usage[prod] += 1
            committed_customers.append(cust_name)

        ws_inv = wb['inventory']
        hi = [ws_inv.cell(1, c).value for c in range(1, ws_inv.max_column + 1)]
        n_col = hi.index('name') + 1
        s_col = hi.index('stock_qty') + 1
        for row in ws_inv.iter_rows(min_row=2):
            name = row[n_col-1].value
            if name in usage:
                old = row[s_col-1].value or 0
                row[s_col-1].value = max(0, old - usage[name])

        wb.save(tmp_path)

        from flask import send_file
        return send_file(tmp_path, as_attachment=True,
                          download_name='snatched_beauty_box_master_updated.xlsx')
    except Exception as e:
        return jsonify({'error': str(e)}), 500


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)
